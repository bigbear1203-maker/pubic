# -*- coding: utf-8 -*-
"""
stock_analysis_log 表頭錯位修復工具
====================================
症狀：用 Excel 打開紀錄檔，發現「程式版本」欄位裡是數字、「實際報酬(%)」
      欄位裡是 3.7 這種明顯放錯位置的值；或用 log_review.py 跑出來的
      統計完全不合理。

成因：紀錄檔的表頭是舊版 analyzer 寫的（欄位較少），但資料列是新版寫的
      （欄位較多）。openpyxl 的 append 會把新版的值一路寫到底，多出來的
      那幾欄沒有表頭，於是從差異點開始，每一欄的值都對到了錯誤的名稱上。
      這種錯位不會噴任何錯誤，但會讓後續所有統計靜靜地算在錯的欄位上。

修復：本工具**不看表頭**，直接讀取資料列，再依目前 analyzer 的正確欄位
      清單重新命名。同時可選擇移除重複紀錄。原檔會先備份。

使用方式：
    python repair_log.py stock_analysis_log_v3.7.xlsx           # 檢查並修復
    python repair_log.py stock_analysis_log_v3.7.xlsx --check   # 只檢查不改
    python repair_log.py stock_analysis_log_v3.7.xlsx --dedup   # 順便移除重複列
"""

from __future__ import annotations

import argparse
import importlib.util
import shutil
import sys
from pathlib import Path

import pandas as pd

SHEET = "分析紀錄"
SCRIPT_DIR = Path(__file__).resolve().parent


def find_analyzer() -> Path | None:
    for name in ("claude_stock_analyzer_v3.7.py", "claude_stock_analyzer_v3.6.py"):
        for base in (SCRIPT_DIR, SCRIPT_DIR.parent, SCRIPT_DIR / "tools"):
            p = base / name
            if p.exists():
                return p
    return None


def load_schema() -> list[str]:
    p = find_analyzer()
    if p is None:
        print("✗ 找不到 claude_stock_analyzer_v3.7.py，無法取得正確的欄位清單。")
        print("  請把本工具放在與 analyzer 同一個資料夾（或其 tools/ 子目錄）。")
        sys.exit(1)
    spec = importlib.util.spec_from_file_location("analyzer", p)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    print(f"[結構] 取自 {p.name}：{len(m.EXCEL_LOG_COLUMNS)} 個欄位")
    return list(m.EXCEL_LOG_COLUMNS)


def inspect(path: Path, schema: list[str]) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {"ok": False, "reason": "檔案是空的"}

    header = list(rows[0])
    data = [list(r) for r in rows[1:]]
    # 去掉表頭尾端的空白欄位再比較，避免 Excel 存檔留下的空欄造成誤判
    trimmed = list(header)
    while trimmed and trimmed[-1] in (None, ""):
        trimmed.pop()

    widths = {len(r) for r in data} if data else set()
    return {
        "ok": True,
        "header": header,
        "trimmed_header": trimmed,
        "data": data,
        "n_rows": len(data),
        "header_width": len(trimmed),
        "data_widths": widths,
        "schema_width": len(schema),
        "aligned": trimmed == schema,
    }


def diagnose(info: dict, schema: list[str]) -> bool:
    """回傳 True 代表需要修復。"""
    print(f"\n[檢查] 資料列數：{info['n_rows']}")
    print(f"       表頭欄位數（去除尾端空白）：{info['header_width']}")
    print(f"       資料列欄位數：{sorted(info['data_widths'])}")
    print(f"       正確欄位數：{info['schema_width']}")

    # 表頭對、寬度也對，不代表資料沒錯位。混版 append 的檔案讀出來
    # 每一列都會被補齊到相同寬度，光看寬度看不出問題——必須檢查內容。
    bad_rows = []
    for i, r in enumerate(info["data"]):
        rr = list(r)[:info["schema_width"]]
        rr += [None] * (info["schema_width"] - len(rr))
        if _row_score(rr, schema) < 4:
            bad_rows.append(i + 2)      # +2：Excel 列號從 1 起算且第 1 列是表頭

    if info["aligned"] and info["data_widths"] <= {info["schema_width"]} and not bad_rows:
        print("\n✓ 表頭與資料一致，不需要修復。")
        return False

    print("\n✗ 偵測到錯位：")
    if bad_rows:
        print(f"    有 {len(bad_rows)} 列的內容與欄位名稱對不上"
              f"（Excel 列號：{bad_rows[:12]}{' …' if len(bad_rows) > 12 else ''}）")
        sample = list(info["data"][bad_rows[0] - 2])
        idx = schema.index("程式版本")
        if idx < len(sample):
            print(f"    例如第 {bad_rows[0]} 列的「程式版本」欄位是 {sample[idx]!r}，"
                  f"但那應該是版本號 3.7")
        print(f"    成因：升版之後繼續往舊檔 append，新舊版本的資料列混在同一個檔案。")
    if not info["aligned"]:
        for i, (h, s) in enumerate(zip(info["trimmed_header"], schema)):
            if h != s:
                print(f"    第 {i + 1} 欄開始不一致：")
                print(f"      檔案表頭：{h}")
                print(f"      應該是：  {s}")
                break
        else:
            print(f"    表頭長度 {info['header_width']} ≠ 正確長度 {info['schema_width']}")
    if info["data_widths"] - {info["schema_width"]}:
        print(f"    有資料列的欄位數不等於 {info['schema_width']}：{sorted(info['data_widths'])}")
    print("\n  影響：從不一致的那一欄開始，之後每一欄的值都對到錯誤的名稱上。")
    print("  這不會噴錯，但 log_review.py 等工具會靜靜地拿錯欄位做統計。")
    return True


# 歷史欄位結構。新版若在「中間」插入欄位，舊資料就不能靠尾端補空來對齊——
# 那樣會讓插入點之後的每一欄再次錯位（本工具原本就踩過這個坑）。這裡用
# 「欄位寬度 → 當時的欄位清單」反推舊資料的真實結構，再依欄位名稱對映到
# 目前的結構，缺的補空、多的丟棄。
#
# 由目前結構往回推導，不需手動維護整份清單：
#   106 欄：目前（新增 操作建議／建議買進價／建議理由）
#   103 欄：新增 較同批次落後天數／是否較同批次落後 之後
#   101 欄：v3.7 初版
COLUMNS_ADDED_LATER = [
    ["操作建議", "建議買進價", "建議理由"],          # 106 → 103
    ["較同批次落後天數", "是否較同批次落後"],          # 103 → 101
]


def historical_schemas(current: list[str]) -> dict:
    """回傳 {欄位寬度: 該版本的欄位清單}，含目前版本與所有已知舊版。"""
    out = {len(current): list(current)}
    cols = list(current)
    for removed in COLUMNS_ADDED_LATER:
        cols = [c for c in cols if c not in removed]
        out.setdefault(len(cols), list(cols))
    return out


# 用來辨識「這一列是哪個版本寫的」的特徵欄位。
# 不能只看寬度：openpyxl 讀取時會把所有列補齊到工作表的最大寬度，
# 舊版列（值較少）與新版列（值較多）讀出來一樣寬，光看寬度分不出來。
# 而同一個檔案裡混著兩種版本的列，是升版後繼續 append 舊檔的必然結果。
def _row_score(row: list, cols: list[str]) -> int:
    """把一列套上某個欄位結構，檢查幾個特徵欄位是否合理。分數越高越可能是對的。"""
    idx = {c: i for i, c in enumerate(cols)}

    def val(name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    score = 0
    v = val("程式版本")
    if isinstance(v, (int, float)) and not isinstance(v, bool) and 1 <= float(v) <= 20:
        score += 3                      # 版本號一定是 3.x，最強的特徵
    elif v is None:
        score += 0
    else:
        score -= 3                      # 是字串就一定錯位了

    v = val("model_quality")
    if v in (None, "not_reliable", "weak", "usable_with_caution", "unknown"):
        score += 2
    else:
        score -= 2

    v = val("三大法人連續方向")
    if v in (None, "買超", "賣超"):
        score += 1
    else:
        score -= 1

    v = val("實際方向")
    if v in (None, "漲", "跌", "平"):
        score += 1
    else:
        score -= 1

    v = val("綜合結論")
    if v is None or isinstance(v, str):
        score += 1
    else:
        score -= 1
    return score


def repair(path: Path, info: dict, schema: list[str], dedup: bool) -> None:
    w = len(schema)
    known = historical_schemas(schema)
    dropped = 0

    # 逐列判斷版本。同一個檔案完全可能同時存在多種版本的列——
    # 升版之後繼續往舊檔 append 就會這樣，而且不會有任何錯誤訊息。
    # 使用者實際遇到的症狀是「程式版本」欄位出現 [3.7, '漲', '跌']：
    # 3.7 是新版列，'漲'/'跌' 是舊版列錯位 3 欄後跑進來的「實際方向」。
    candidates = [known[k] for k in sorted(known, reverse=True)]
    rows_by_schema: dict[int, int] = {}
    fixed = []

    for r in info["data"]:
        r = list(r)
        best, best_score = None, None
        for cand in candidates:
            rr = r[:len(cand)] if len(r) > len(cand) else r + [None] * (len(cand) - len(r))
            sc = _row_score(rr, cand)
            if best_score is None or sc > best_score:
                best, best_score, best_row = cand, sc, rr
        rows_by_schema[len(best)] = rows_by_schema.get(len(best), 0) + 1
        # 先用該列自己的結構建 Series，再依欄位名稱對映到目前結構
        fixed.append(dict(zip(best, best_row)))

    if len(rows_by_schema) > 1:
        print(f"\n  ⚠ 這個檔案裡混著多種欄位版本的資料列：")
        for wdt, n in sorted(rows_by_schema.items(), reverse=True):
            missing = [c for c in schema if c not in known.get(wdt, schema)]
            tag = "目前版本" if wdt == w else f"舊版，缺 {missing}"
            print(f"      {wdt} 欄 × {n} 列（{tag}）")
        print(f"  成因：升版之後繼續往舊檔 append。每一列會依自己的版本分別對齊。")
    elif rows_by_schema and list(rows_by_schema)[0] != w:
        wdt = list(rows_by_schema)[0]
        missing = [c for c in schema if c not in known.get(wdt, schema)]
        print(f"\n  資料是 {wdt} 欄版本寫的，目前結構是 {w} 欄。")
        print(f"  依欄位名稱對映，新增的欄位留空：{missing}")

    # 依欄位名稱建表：中間插入的新欄位會落到正確位置並留空，
    # 而不是把舊資料整段往後推。
    df = pd.DataFrame(fixed).reindex(columns=schema)

    if dedup:
        before = len(df)
        # 同一檔、同一資料基準日只保留最後一筆（與 log_review 的去重規則一致）
        if {"股票代碼", "股價日期(資料基準日)", "執行時間"} <= set(df.columns):
            df = (df.sort_values("執行時間")
                    .groupby(["股票代碼", "股價日期(資料基準日)"], as_index=False)
                    .last())
            df = df[schema]
        dropped = before - len(df)

    backup = path.with_name(path.stem + "_修復前備份" + path.suffix)
    shutil.copy2(path, backup)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=SHEET, index=False)

    print(f"\n✓ 已修復：{path}")
    print(f"  原檔備份：{backup.name}")
    print(f"  資料列數：{len(df)}" + (f"（移除 {dropped} 筆重複）" if dropped else ""))

    # 修復後的抽樣驗證
    print("\n[驗證] 隨機抽查幾個容易看出錯位的欄位：")
    for col in ["程式版本", "綜合結論", "三大法人連續方向", "model_quality"]:
        if col in df.columns:
            vals = df[col].dropna().unique()[:3]
            print(f"    {col}: {list(vals)}")


# 已知「本來就長得像、但刻意分開」的欄位配對。這些不是缺陷，列出來是
# 為了讓稽核報告不要每次都把它們當成問題。
BY_DESIGN_PAIRS = {
    ("是否盤中執行", "已剔除未完成K棒"):
        "前者是執行當下在不在盤中，後者是實際上有沒有砍掉未完成K棒。"
        "09:10 執行時可能盤中=True 但沒東西可剔除=False，資料其實乾淨。",
    ("較同批次落後天數", "是否較同批次落後"):
        "數值與布林旗標並存是刻意的——Excel 篩選布林欄比寫數值條件方便得多。",
    ("資料落後天數", "資料是否停滯"):
        "同上；停滯與否的門檻是設定值，把判斷結果一併存下來才不用回頭推算。",
    ("殖利率(%)", "殖利率是否異常"): "數值與異常旗標並存，方便直接篩掉異常值。",
    ("負債權益比", "負債權益比是否異常"): "同上。",
}


def audit_columns(path: Path, schema: list[str]) -> int:
    """
    稽核欄位是否有重複或無用。

    ⚠ 判讀重點：「這批資料每一列的值都相同」不等於「這兩欄是重複的」。
    資料乾淨時，一堆布林旗標會通通是 False，彼此當然相同——那是巧合，
    不是缺陷。所以下面會把「已知刻意分開的配對」單獨標示出來，
    剩下的才需要你判斷。
    """
    df = pd.read_excel(path, sheet_name=SHEET)
    work = df.drop(columns=[c for c in ("完整終端輸出",) if c in df.columns])
    n = len(work)

    print("\n" + "=" * 66)
    print(f"  欄位稽核：{path.name}（{n} 列 × {len(df.columns)} 欄）")
    print("=" * 66)

    names = list(schema)
    dup_names = sorted({c for c in names if names.count(c) > 1})
    print(f"\n【1】名稱重複的欄位：{dup_names if dup_names else '無'}")

    print(f"\n【2】這批資料中值完全相同的欄位配對")
    import itertools
    pairs = []
    for x, y in itertools.combinations(work.columns, 2):
        sx, sy = work[x], work[y]
        if sx.isna().all() and sy.isna().all():
            continue
        try:
            if sx.equals(sy):
                pairs.append((x, y))
        except Exception:
            pass
    if not pairs:
        print("    無")
    else:
        known, unknown = [], []
        for x, y in pairs:
            note = BY_DESIGN_PAIRS.get((x, y)) or BY_DESIGN_PAIRS.get((y, x))
            (known if note else unknown).append((x, y, note))
        if known:
            print("\n    ── 刻意分開的（不是缺陷）──")
            for x, y, note in known:
                print(f"    · {x} / {y}")
                print(f"      {note}")
        if unknown:
            print(f"\n    ── 需要你判斷的 {len(unknown)} 組 ──")
            for x, y, _ in unknown:
                print(f"    · {x} / {y}")
            print(f"\n    ⚠ 只有 {n} 列資料時，布林旗標很容易剛好都相同。")
            print(f"      累積到數十列、且出現過異常情況之後再看這一段才準。")

    print(f"\n【3】整欄全空的欄位")
    empty = [c for c in work.columns if work[c].isna().all()]
    # 空欄不一定是問題。分成三類：本來就該空的、有對應指令可以補的、真的異常。
    expected = {
        "跳過原因": "沒有標的被跳過（好事）",
        "交易日曆提醒": "假日清單尚未到期",
        "執行當下價格": "沒有盤中執行（只有盤中才會記錄）",
        "Strategy_Decision_信心(%)": "沒有出現 Buy/Sell 訊號時本來就是空的",
        "實際目標日收盤": "等預測目標日收盤後由 log_review 回填",
        "實際報酬(%)": "同上", "實際方向": "同上",
        "是否命中_邏輯迴歸": "同上", "是否命中_RF": "同上",
        "是否命中_綜合分數": "同上", "回填時間": "同上",
        "外資佔量比重5日斜率": "本地快取需累積 5 個不同交易日才會開始計算",
    }
    fixable = {
        "操作建議": "advice", "建議買進價": "advice", "建議理由": "advice",
        "建議停損價": "advice", "建議目標價": "advice", "Risk_Reward_Ratio": "advice",
    }
    normal = [c for c in empty if c in expected]
    todo = [c for c in empty if c in fixable]
    odd = [c for c in empty if c not in expected and c not in fixable]

    if normal:
        print(f"    · 本來就該空的 {len(normal)} 欄：")
        for c in normal:
            print(f"      {c}：{expected[c]}")
    if todo:
        print(f"\n    → 可以補起來的 {len(todo)} 欄：")
        print(f"      {'、'.join(todo)}")
        print(f"      這些是舊版紀錄沒有的欄位，執行下面這行就會補上：")
        print(f"        python stock.py advice")
    if odd:
        print(f"\n    ⚠ 預期不該空、需要你確認的 {len(odd)} 欄：")
        for c in odd:
            print(f"      · {c}")
    if not empty:
        print("    無")

    print("\n" + "=" * 66)
    return 0


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="stock_analysis_log 表頭錯位修復")
    ap.add_argument("log", type=Path)
    ap.add_argument("--check", action="store_true", help="只檢查，不修改檔案")
    ap.add_argument("--dedup", action="store_true",
                    help="同時移除重複紀錄（同一檔、同一資料基準日只留最後一筆）")
    ap.add_argument("--audit", action="store_true",
                    help="稽核欄位是否有重複或無用（不修改檔案）")
    args = ap.parse_args(argv)

    if not args.log.exists():
        print(f"找不到檔案：{args.log}")
        return 1

    schema = load_schema()
    if args.audit:
        return audit_columns(args.log, schema)

    info = inspect(args.log, schema)
    if not info["ok"]:
        print(f"✗ {info['reason']}")
        return 1

    need = diagnose(info, schema)
    if not need and not args.dedup:
        return 0
    if args.check:
        print("\n（--check 模式，未修改檔案）")
        return 0

    repair(args.log, info, schema, args.dedup)
    print("\n修復後請重新執行：python log_review.py <紀錄檔> --write-back")
    return 0


if __name__ == "__main__":
    sys.exit(main())
