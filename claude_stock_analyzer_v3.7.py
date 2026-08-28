"""
股票基本面 + 技術面 + 籌碼面整合分析工具
==========================================
支援台股(需加 .TW 或 .TWO 後綴)與美股代碼。
資料來源: Yahoo Finance (yfinance) + 台灣證交所/櫃買中心公開資訊(台股籌碼面)

版本: v3.7
版本紀錄:
    v3.7 - 依 80 筆長期紀錄(2026-08-26~08-28)的實證診斷修正。診斷全文見
           docs/程式診斷報告.md，每一項修正都對應紀錄中可重現的證據。
           [重大修正] 盤中執行使用未完成K棒：analyze() 先前直接使用
                    yfinance 回傳的最後一根K棒，盤中執行時該K棒尚未收盤，
                    Close 其實是「當下成交價」，而隔日模型的核心特徵
                    intraday_return=(Close-Open)/Open 訓練時看到的全是
                    完整交易日，盤中餵進去的是「開盤到現在」，分布不同。
                    實測：2891C.TW 於 2026-08-28 13:15(盤中)算出上漲機率
                    73.9%，同日 15:01(收盤後)以同一「資料基準日」算出
                    34.8%，差 39.2 個百分點且方向相反；整份紀錄唯一一次
                    Sell 訊號即由此產生。本版新增 _trim_incomplete_bar()，
                    盤中執行時剔除未完成K棒，以前一完整交易日為基準。
                    註：今日(Open→Close)模型不受此限——它的特徵只用今天
                    開盤價與昨日(含)以前的值，從不碰今天的Close，因此
                    仍使用含當日K棒的完整序列，僅排除當日進入訓練集。
           [重大修正] 今日模型的標籤洩漏：intraday_close_probability_
                    walkforward() 最終訓練用的 X_all[-train_window:] 包含
                    「今天」這一列，而該列的 target 正是今天的 Open→Close
                    答案，等於讓模型先看過答案再預測同一列。隔日模型因
                    target=shift(-1)被 dropna() 砍掉末列而沒有這個問題，
                    兩者不一致。本版改為明確排除被預測列後再訓練。
           [重大修正] 資料停滯不報錯：00684R.TW 於 2026-08-28 連跑三次，
                    資料基準日全部停在 08-26、股價全部 15.25 未變動，
                    程式仍照常輸出機率與「中性偏多」結論。新增
                    _check_price_freshness()，資料落後超過容許天數即
                    標記並拒絕輸出交易決策。
           [重大修正] 「資料不足」與「分析後認為中性」無法區分：
                    020020.TW 連跑三次，市場狀態 unknown、ADX 與兩個模型
                    機率全為 NaN，卻仍給綜合分數 0、結論「中性（訊號混合）」。
                    新增 _should_analyze() 前置把關，不合格者記為
                    Skipped_資料不足 / Skipped_資料停滯，不再混入中性。
           [策略層] 模型品質關卡改用 Wilson 信賴區間下限：原本的
                    「樣本外準確率 < 55% 就封鎖」沒有考慮樣本數，
                    小樣本的高準確率會直接通關。改為要求信賴區間下限
                    高於 50%（n=60 需 ≥62.7%、n=250 需 ≥56.2%、
                    n=1000 需 ≥53.1%），樣本越少門檻越嚴。
           [策略層] 新增 EV_Decision 影子決策：以「扣掉台股來回成本
                    (手續費6折×2 + 證交稅0.3% ≈ 0.47%)後期望值是否為正」
                    判斷方向，與既有的雙模型一致性 Strategy_Decision
                    並行記錄、互不干擾。刻意採影子模式——在累積足夠
                    樣本證明哪一種決策規則較好之前，不片面改掉你原本
                    的交易邏輯，只把比較所需的資料一起記下來。
           [修正] next_trading_day_estimate() 補上台股國定假日：原本只
                    跳週末，遇連假會把「預測目標日」標錯，事後命中率
                    統計會對錯日子、靜靜汙染整份驗證結果。假日清單
                    維護到期時會明確提示，不會默默算出錯誤日期。
           [Excel] 新增欄位：各模型 walk-forward 樣本數（先前 walk-forward
                    有算出 sample_size 卻沒有記錄，導致無法判斷任何一筆
                    準確率是幾筆樣本算出來的）、準確率信賴區間下限、
                    是否盤中執行、已剔除未完成K棒、執行當下價格、
                    資料落後天數、資料是否停滯、今日模型是否已知結果、
                    EV_Decision 與淨期望值、假日曆提醒，以及供
                    tools/log_review.py 事後回填的實際結果欄位
                    （實際目標日收盤／實際報酬(%)／是否命中_*／回填時間）。
                    欄位結構變動，沿用既有保護機制自動另存新檔。
           [健壯性] _append_to_excel_log() 寫入失敗時改存 CSV 備援，
                    不再讓該筆分析結果直接遺失。
    v3.6 - 依使用者需求文件調整(短期+可行中期項目)：
           [策略層] 新增「隔日 Strategy_Decision」：雙模型(邏輯迴歸+RF)方向一致
                    且雙方信心都達門檻(預設58%)才輸出 Buy/Sell，否則一律 Wait，
                    避免模型意見分歧或信心不足時仍發出明確訊號。
           [策略層] 新增 Risk_Reward_Ratio：僅在有明確 Buy/Sell 訊號時，用
                    進場價 ± 1.5×ATR 估算停損、2×ATR 估算目標價，算出風報比。
                    這是粗略的風險參考，不是精確的獲利目標。
           [策略層] 新增期望值(EV)：在隔日方向 walk-forward 回測中，額外記錄
                    「若照當時模型判斷方向交易」的隔日報酬，算出勝率/平均獲利/
                    平均虧損/EV，取代單純看準確率。
           [監控] 新增模型衰退監控：取 walk-forward 回測中最近30個交易日的
                    樣本外準確率，低於50%時印出警報建議暫停/重新訓練。此數字
                    來自當次重新執行的 walk-forward 回測，不是從 Excel 歷史
                    log 反推(避免 log 缺筆、非交易日、同日多次執行造成的
                    對齊問題，改用當次回測的最後30筆更穩健)。
           [特徵] 新增市場狀態特徵：ADX 判斷的趨勢市/盤整市 one-hot 編碼，
                    餵給方向機率模型(先前只用來調整技術面規則式評分，沒有讓
                    機率模型知道現在處於什麼市場環境)。
           [特徵] 新增大盤關聯指標特徵：費城半導體指數(^SOX)、美元兌台幣
                    (TWD=X)、美國10年期公債殖利率(^TNX)的乖離率與5日變動率，
                    同一次執行(含批次查詢多檔)只抓取一次、共用快取。任一指標
                    抓取失敗只會少那組特徵，不影響其餘功能。
           [籌碼面] 外資/投信/自營合計買賣超新增「連續同方向天數」(在近期
                    抓取到的交易日窗格內)；新增「外資買賣超佔當日成交量比重」
                    (與 hist 當日成交量比對)。
           [Excel] 紀錄檔新增對應欄位；欄位結構變動，沿用舊版保護機制，
                    第一次執行會自動另存新檔，不覆蓋 v3.5 舊檔案。
           [未做，僅研究] 借券賣出餘額變化：已確認 TWSE 有對應端點
                    (exchangeReport/TWT93U)，抓取模式與現有籌碼快取一致，
                    可行但本版未實作。開盤後5分鐘成交量：yfinance 對台股
                    分鐘級資料不穩定，可行性低，本版未實作、亦未規劃。
           [未做，僅規劃] 波動率回歸模型(策略層面從預測方向改為預測波動)：
                    屬於較大架構調整，本版未實作，待獨立討論。
           [同版本後續追加] 特徵重要性：隔日模型最終訓練額外保留
                    feature_importances_(RF)/標準化係數絕對值(邏輯迴歸，
                    非嚴謹重要性)，畫面與 Excel 印出 Top5。
           [同版本後續追加] 產業標註：sector/industry 存進 metrics/Excel；
                    半導體相關產業明確標註 SOX 特徵為「產業週期代理指標」，
                    其他產業目前無對應代理指標，不硬造權重。
           [同版本後續追加] 籌碼集中度(外資買賣超/當日成交量)升級為本地
                    快取時間序列，仿照外資5日斜率算 5 日斜率——這仍是報告
                    層級指標，TWSE 免費端點無歷史資料可回填，無法真正變成
                    模型訓練特徵，跟外資5日斜率同樣的天花板。
           [同版本後續追加] 模型品質雙重防護：任一模型樣本外準確率低於55%
                    時，Strategy_Decision 強制 Wait(不論58%信心門檻是否
                    達標)；新增機器可讀 model_quality 標籤
                    (not_reliable/weak/usable_with_caution/unknown)。
           [同版本後續追加，code review 修正] calc_kd() 修正除以0風險
                    (high_max==low_min 時)；美股路徑 fundamental_summary
                    抓到的 info 傳給 chip_summary_us 避免重抓；殖利率/
                    負債權益比新增負值異常檢查；chip_summary_us 內部人
                    交易關鍵字新增 acquired/disposed，並新增「未分類」
                    計數；_run_batch() 過濾空白代碼、印出成功/失敗統計。
    v3.5 - 新增「完整終端輸出」記錄：分析過程中畫面上印出的所有文字
           (基本面/技術面/籌碼面明細、方向機率、方法論限制等全部內容)
           會完整存進 Excel 紀錄檔新增的「完整終端輸出」欄位，不只存
           結構化數值。若單次輸出超過約 30,000 字元(Excel 儲存格上限
           約 32,767 字元)，會截斷並標註提示，避免寫入失敗；畫面上的
           完整輸出不受影響。因欄位結構變動，沿用舊版「欄位不一致就
           另開新檔」的保護機制，第一次執行會自動另存新的紀錄檔，不會
           覆蓋 v3.3/v3.4 產生的舊檔案。
           [修正] _TeeOutput 補上 __getattr__，把未實作的屬性(如
           .isatty()/.encoding)轉發給真正的 sys.stdout，避免第三方
           套件檢查這些屬性時，因為屬性不存在而中途噴錯、來不及寫入
           Excel。
    v3.4 - 新增「今日收盤方向機率」(開盤後估計，用隔夜跳空+前一交易日已知
           特徵預測當天Open→Close方向，嚴格避免用到當天收盤資訊造成資料
           洩漏)；「隔日方向機率」區塊加上明確日期標示(資料基準日/預測
           目標日，估計下一交易日僅跳過六日未過濾國定假日)；Excel紀錄檔
           欄位同步更新(股價日期改名為「股價日期(資料基準日)」、新增
           「預測目標日」欄位、今日/隔日模型結果分開存放)。
    v3.3 - 新增長期記錄功能：每次分析結果自動新增一列到固定 Excel 檔案
           (stock_analysis_log.xlsx，與程式同目錄)，供長期追蹤比對。
           fundamental_summary/chip_summary 系列函式回傳值新增 metrics
           結構化字典(供 Excel 匯出使用，不再從畫面文字反推數字)。
    v3.2 - 新增多檔股票查詢支援(command line可傳入多個代碼，逐檔分析、
           個別失敗不中斷、檔與檔之間加入延遲降低TWSE請求壓力、結尾印出
           批次結果總覽)。
    v3.1 - 加入 ADX 市場狀態過濾器(趨勢市/盤整市動態調整技術面邏輯)；
           重構方向機率預估特徵工程(overnight_gap/intraday_return/波動度/
           指標變化率)；新增 Random Forest 對照組(選用，需 scikit-learn)；
           新增外資買賣超5日斜率(本地快取，逐日累積)。
    v3   - 使用者上傳的基準版本。

使用前請先安裝套件:
    pip install yfinance pandas numpy requests openpyxl
    pip install scikit-learn   # 選用，啟用 Random Forest 對照組時才需要

使用方式:（把 <檔名> 換成你實際存的檔名，例如 claude_stock_analyzer_v3.3.py）
    python <檔名> 2330.TW
    python <檔名> AAPL
    python <檔名> 2330.TW 2317.TW AAPL   # 多檔查詢

長期記錄:
    每次執行都會在程式所在目錄自動新增/更新 stock_analysis_log.xlsx，
    把這次分析的關鍵數值(股價、三面向分數、波動區間、方向機率等)新增一列。
    不會覆蓋舊紀錄；同一天執行多次會累積多列。

注意: 本工具僅整理歷史財報、技術指標與籌碼數據作為參考,
不構成投資建議,也不提供任何形式的股價預測——短期股價
走勢受消息面、市場情緒等無法量化的因素影響,任何模型都無法
準確預測未來走勢。文末提供的「隔日歷史經驗波動區間」是根據過去報酬分位數計算的
風險參考範圍,並非預測值,也不代表未來有固定機率落在區間內。
"""

import sys
import os
import io
import time
import datetime
import numpy as np
import pandas as pd

try:
    import yfinance as yf
except ImportError:
    print("請先安裝 yfinance: pip install yfinance")
    sys.exit(1)

try:
    import requests
except ImportError:
    requests = None

try:
    from openpyxl import Workbook, load_workbook
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ----------------------------
# 資料品質防護（v3.7 新增）
# ----------------------------
#
# 這一整段是 v3.7 依 80 筆長期紀錄的實證診斷新增的把關機制。每個函式的
# docstring 都附上紀錄中對應的實際證據，方便日後判斷這道關卡還需不需要。

# 台股連續交易時段
TW_SESSION_OPEN = datetime.time(9, 0)
TW_SESSION_CLOSE = datetime.time(13, 30)
# 收盤後緩衝：資料源更新需要時間，太早抓仍可能拿到未定案的收盤價
TW_DATA_SETTLE = datetime.time(14, 0)

# 台股來回交易成本：手續費 0.1425% × 折扣 × 買賣兩次 + 賣出證交稅 0.3%
FEE_DISCOUNT = 0.6          # 券商手續費折扣，請依你實際使用的券商調整
SECURITIES_TAX_PCT = 0.3    # 現股賣出證交稅；當沖減半請改 0.15


def round_trip_cost_pct(fee_discount=FEE_DISCOUNT, tax_pct=SECURITIES_TAX_PCT):
    """一買一賣的總成本百分比。這是任何策略必須先跨過的門檻。"""
    return 0.1425 * fee_discount * 2 + tax_pct


def _is_tw_symbol(symbol):
    return str(symbol).upper().endswith((".TW", ".TWO"))


def is_intraday_now(symbol, now=None):
    """
    現在是否處於該市場的盤中時段。台股用本機時間判斷（假設你在台灣執行）。
    美股因為時區換算牽涉夏令時間，這裡不猜，一律回傳 False，改由
    _trim_incomplete_bar() 用「最後一根K棒是不是今天」搭配判斷。
    """
    now = now or datetime.datetime.now()
    if not _is_tw_symbol(symbol):
        return False
    return TW_SESSION_OPEN <= now.time() < TW_DATA_SETTLE


def _trim_incomplete_bar(hist, symbol, now=None):
    """
    ⚠ v3.7 最重要的修正。

    問題：yfinance 的 history() 在盤中會把「今天這根還沒收盤的K棒」一起
    回傳，此時 Close 其實是當下成交價，不是收盤價。而隔日模型的核心特徵
        intraday_return = (Close - Open) / Open
    訓練時看到的全是「完整交易日的開盤→收盤」，盤中拿到的卻是
    「開盤→現在」——輸入分布不同，模型等於在沒見過的資料上做外插。

    實測證據（同一檔、同一個資料基準日，只差在執行時間）：
        2891C.TW 2026-08-28  13:15 → 73.9%   15:01 → 34.8%   (差 39.2pp，方向相反)
        2609.TW  2026-08-27  11:33 → 49.2%   15:03 → 66.7%   (差 17.5pp，方向相反)
        3037.TW  2026-08-28  13:25 → 54.5%   15:11 → 65.8%   (差 11.3pp)
    80 筆紀錄中有 30 筆是盤中執行的，全部受影響；整份紀錄唯一一次
    Sell 訊號就是這樣產生的。

    處理：盤中執行時剔除未完成K棒，以前一個完整交易日為基準。這會讓
    盤中執行的結果等同於「昨天收盤後跑的結果」——這是誠實的做法：
    盤中本來就還沒有今天的收盤資訊，硬算不會讓資訊變多，只會讓錯誤變隱形。

    回傳 (處理後的 hist, 是否有剔除, 被剔除那根K棒的收盤價或 None)。
    """
    if hist is None or hist.empty:
        return hist, False, None

    now = now or datetime.datetime.now()
    last_date = hist.index[-1].date()

    if last_date == now.date() and is_intraday_now(symbol, now):
        live_price = float(hist["Close"].iloc[-1])
        return hist.iloc[:-1].copy(), True, live_price
    return hist, False, None


# 同一次批次執行中，各標的最新交易日的最大值。用來偵測「這一檔比同批
# 其他檔落後一個交易日」這種相對落後——它不會觸發絕對天數門檻，卻是
# 資料源沒更新的典型徵兆。由 _run_batch() 在批次開始時重置。
_BATCH_LATEST_DATE = None


def _check_batch_lag(price_date):
    """
    與同批次其他標的比較，判斷這一檔的資料是不是落後。

    實測證據：00684R.TW 在 2026-08-28 的三次執行中，資料基準日都停在
    2026-08-26，而同一批的其他標的都是 2026-08-27——落後整整一個交易日。
    但它只落後 2 個日曆日，低於 _check_price_freshness() 的 4 天門檻，
    所以絕對天數的檢查抓不到。相對比較才抓得到這種情況。

    回傳 (是否落後, 落後天數, 訊息)。批次中第一檔沒有比較基準，一律回報未落後。
    """
    global _BATCH_LATEST_DATE
    if price_date is None:
        return False, None, ""
    if _BATCH_LATEST_DATE is None or price_date > _BATCH_LATEST_DATE:
        _BATCH_LATEST_DATE = price_date
        return False, 0, ""
    lag = (_BATCH_LATEST_DATE - price_date).days
    if lag <= 0:
        return False, 0, ""
    return True, lag, (
        f"本檔資料基準日 {price_date} 比同批次其他標的的最新日 "
        f"{_BATCH_LATEST_DATE} 落後 {lag} 天，該檔資料源可能未更新，"
        f"分析結果請降低採信程度"
    )


def _check_price_freshness(hist, max_lag_days=4, now=None):
    """
    價格資料是否停滯。

    這道檢查針對的是「絕對落後」：下市、停牌、代碼變更這類資料源整個
    斷掉的情況。max_lag_days=4 是為了涵蓋週末（週五收盤 → 週二執行 = 4 天）。

    ⚠ 它抓不到「只落後一個交易日」的情況。實測證據：00684R.TW 於
    2026-08-28 跑了三次，資料基準日全部停在 2026-08-26（同批次其他標的
    都是 2026-08-27），但那只落後 2 個日曆日，低於本門檻。這一類要靠
    _check_batch_lag() 的相對比較才抓得到——兩道檢查互補，缺一不可。
    """
    now = now or datetime.datetime.now()
    if hist is None or hist.empty:
        return {"ok": False, "stale": True, "lag_days": None, "last_date": None,
                "message": "沒有任何價格資料"}

    last_date = hist.index[-1].date()
    lag = (now.date() - last_date).days
    stale = lag > max_lag_days
    return {
        "ok": not stale, "stale": stale, "lag_days": lag, "last_date": last_date,
        "message": (f"價格資料停滯：最新交易日 {last_date}，距今 {lag} 天，"
                    f"超過容許的 {max_lag_days} 天，本次結果不應採用"
                    if stale else f"價格資料為 {last_date}（距今 {lag} 天）"),
    }


def _should_analyze(hist, min_bars=300):
    """
    在跑完整分析之前，先判斷這檔標的的資料夠不夠格被分析。

    實測證據：020020.TW 連跑三次，市場狀態 unknown、ADX 與兩個模型機率
    全部是 NaN，卻仍然寫入紀錄、綜合分數給 0、結論寫「中性（訊號混合，
    各面向未形成一致方向）」。讀報表的人會以為系統看過它了，實際上系統
    根本沒有資料可看。「資料不足」跟「分析後認為中性」是兩件完全不同的事，
    不該長得一樣。
    """
    if hist is None or hist.empty:
        return {"ok": False, "reason": "查無價格資料"}
    if len(hist) < min_bars:
        return {"ok": False,
                "reason": f"歷史資料僅 {len(hist)} 個交易日，少於模型所需的 {min_bars} 日"}
    if hist["Close"].tail(60).nunique() <= 1:
        return {"ok": False, "reason": "近 60 日收盤價無變動，可能是流動性極低或資料異常"}
    if "Volume" in hist.columns and float(hist["Volume"].tail(20).median()) <= 0:
        return {"ok": False, "reason": "近 20 日成交量中位數為 0，流動性不足以交易"}
    return {"ok": True, "reason": ""}


def wilson_lower_bound(accuracy_pct, sample_size, z=1.96):
    """
    準確率的 Wilson 信賴區間下限（百分比）。小樣本時比常態近似可靠。
    樣本數不明時回傳 None——「不知道」和「不好」是兩回事，不要混為一談。
    """
    if accuracy_pct is None or not sample_size or sample_size <= 0:
        return None
    p = accuracy_pct / 100.0
    n = int(sample_size)
    d = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / d
    half = z * np.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / d
    return (centre - half) * 100


def assess_accuracy_reliability(accuracy_pct, sample_size, baseline_pct=50.0):
    """
    判斷回報的樣本外準確率是不是真的有優勢，而不是小樣本的隨機波動。

    為什麼要改掉舊的「準確率 < 55% 就封鎖」：那個規則沒有考慮樣本數。
    實測證據：2891C.TW 回報樣本外準確率 74~77%，是全部 19 檔裡最高的，
    也是整份紀錄唯一觸發訊號的一檔；其餘標的的準確率中位數只有 51.4%
    （邏輯迴歸）／52.4%（RF）。這個 74% 到底是真優勢還是小樣本假象，
    在 v3.6 下「無法判斷」——因為 walk-forward 明明算出了 sample_size，
    EXCEL_LOG_COLUMNS 卻沒有把它記下來。

    改用 Wilson 信賴區間下限：只有「連下限都高於基準」才算真有優勢。
    門檻隨樣本數變動：n=60 需 ≥62.7%、n=250 需 ≥56.2%、n=450 需 ≥54.6%、
    n=1000 需 ≥53.1%。樣本越少門檻越嚴，這正是小樣本該有的待遇。
    """
    lower = wilson_lower_bound(accuracy_pct, sample_size)
    if lower is None:
        return {"reliable": False, "label": "unknown", "lower_bound": None,
                "sample_size": sample_size,
                "message": "缺少準確率或樣本數，無法判斷可靠性"}

    reliable = lower > baseline_pct
    if reliable:
        label = "usable_with_caution"
    elif accuracy_pct >= 55:
        label = "weak"          # 點估計看起來不錯，但信賴區間仍蓋到 50%
    else:
        label = "not_reliable"

    return {
        "reliable": reliable, "label": label, "lower_bound": lower,
        "sample_size": int(sample_size),
        "message": (f"準確率 {accuracy_pct:.1f}%（n={int(sample_size)}），"
                    f"95% 信賴區間下限 {lower:.1f}%，"
                    + ("高於 50%，有統計上的優勢跡象"
                       if reliable else f"未高於 {baseline_pct:.0f}%，無法排除只是隨機波動")),
    }


def expectancy_decision(p_up, avg_gain_pct, avg_loss_pct, reliability,
                        min_edge_pct=0.1):
    """
    影子決策（v3.7 新增，不取代 Strategy_Decision）。

    不問「兩個模型是不是都超過 58%」，而是問：照這個機率下注，扣掉
    手續費和證交稅之後，期望值還是正的嗎？
        EV_做多 = p × 平均獲利 − (1−p) × 平均虧損 − 來回成本
        EV_做空 = (1−p) × 平均獲利 − p × 平均虧損 − 來回成本

    為什麼提出這個替代方案：v3.6 的規則是「雙模型方向一致 + 雙方信心
    都 ≥58% + 準確率 ≥55%」。實測 80 筆紀錄裡，兩模型方向一致率只有
    57.1%（機率相關係數 0.47），準確率中位數只有 51~52%，三個條件同時
    成立的機率極低——結果 80 筆裡 79 筆是 Wait。問題不是門檻設太高
    （模型確實沒有可靠優勢），而是這個事實被包裝成「等待訊號」，
    而不是「這套模型目前沒有可交易的優勢」。對投資決策來說，
    這兩句話的意義天差地遠。

    刻意採影子模式：在累積足夠樣本證明哪一種規則較好之前，不片面改掉
    你原本的交易邏輯，只把兩者的判斷都記進 Excel，讓資料自己說話。
    """
    cost = round_trip_cost_pct()
    detail = {"cost_pct": cost, "ev_pct": None, "reason": ""}

    if p_up is None or avg_gain_pct is None or avg_loss_pct is None:
        detail["reason"] = "缺少機率或損益統計，無法計算期望值"
        return "Wait", detail

    if not reliability.get("reliable", False):
        detail["reason"] = "模型準確率信賴區間下限未高於 50%：" + reliability.get("message", "")
        return "Wait", detail

    p = float(p_up)
    ev_long = p * avg_gain_pct - (1 - p) * avg_loss_pct - cost
    ev_short = (1 - p) * avg_gain_pct - p * avg_loss_pct - cost
    detail["ev_long_pct"] = ev_long
    detail["ev_short_pct"] = ev_short
    detail["ev_pct"] = max(ev_long, ev_short)

    if ev_long >= ev_short and ev_long > min_edge_pct:
        detail["reason"] = f"做多期望值 {ev_long:+.2f}%（已扣 {cost:.2f}% 來回成本）"
        return "Buy", detail
    if ev_short > ev_long and ev_short > min_edge_pct:
        detail["reason"] = f"做空期望值 {ev_short:+.2f}%（已扣 {cost:.2f}% 來回成本）"
        return "Sell", detail

    detail["reason"] = (f"最佳方向期望值僅 {max(ev_long, ev_short):+.2f}%，"
                        f"未超過 {min_edge_pct:.2f}% 的最低邊際要求"
                        f"（來回成本 {cost:.2f}%）")
    return "Wait", detail


# ----------------------------
# 技術指標計算函式
# ----------------------------

def calc_ma(df, windows=(5, 20, 60)):
    for w in windows:
        df[f"MA{w}"] = df["Close"].rolling(window=w).mean()
    return df


def calc_rsi(df, period=14):
    """使用 Wilder 常見的 RSI 平滑方式，避免單純 rolling mean 造成訊號跳動。"""
    delta = df["Close"].diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    avg_loss = loss.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    df["RSI"] = 100 - (100 / (1 + rs))
    df.loc[(avg_loss == 0) & (avg_gain > 0), "RSI"] = 100.0
    df.loc[(avg_gain == 0) & (avg_loss > 0), "RSI"] = 0.0
    return df


def calc_macd(df, fast=12, slow=26, signal=9):
    ema_fast = df["Close"].ewm(span=fast, adjust=False).mean()
    ema_slow = df["Close"].ewm(span=slow, adjust=False).mean()
    df["MACD"] = ema_fast - ema_slow
    df["MACD_signal"] = df["MACD"].ewm(span=signal, adjust=False).mean()
    df["MACD_hist"] = df["MACD"] - df["MACD_signal"]
    return df


def calc_kd(df, period=9):
    """
    KD 指標。用 .replace(0, np.nan) 防呆：當 period 天內最高價等於最低價
    (例如連續一字線、極端低量股)，high_max-low_min 會是 0，不處理會產生
    inf/NaN 並污染後續 K/D 值；改為 NaN 後 RSV 也一併 clip 到 0~100，避免
    極端值往下游擴散。
    """
    low_min = df["Low"].rolling(window=period).min()
    high_max = df["High"].rolling(window=period).max()
    price_range = (high_max - low_min).replace(0, np.nan)
    rsv = (df["Close"] - low_min) / price_range * 100
    rsv = rsv.clip(lower=0, upper=100)
    k = rsv.ewm(com=2, adjust=False).mean()
    d = k.ewm(com=2, adjust=False).mean()
    df["K"] = k
    df["D"] = d
    return df


def calc_bollinger(df, window=20, num_std=2):
    mid = df["Close"].rolling(window=window).mean()
    std = df["Close"].rolling(window=window).std()
    df["BB_mid"] = mid
    df["BB_upper"] = mid + num_std * std
    df["BB_lower"] = mid - num_std * std
    return df


def calc_atr(df, period=14):
    """True Range + Wilder smoothing，較接近常見交易軟體的 ATR 定義。"""
    high_low = df["High"] - df["Low"]
    high_close = (df["High"] - df["Close"].shift()).abs()
    low_close = (df["Low"] - df["Close"].shift()).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    df["TR"] = tr
    df["ATR"] = tr.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    return df


def calc_adx(df, period=14):
    """
    Wilder ADX：判斷「趨勢強度」而非方向。
    ADX >= 25 通常視為有明確趨勢；ADX < 20 通常視為盤整（無方向）；20~25 為過渡帶。
    這是市場狀態過濾器(Market Regime Filter)的依據，不是買賣訊號本身。
    """
    up_move = df["High"].diff()
    down_move = -df["Low"].diff()
    plus_dm = np.where((up_move > down_move) & (up_move > 0), up_move, 0.0)
    minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0.0)

    high_low = df["High"] - df["Low"]
    high_close = (df["High"] - df["Close"].shift()).abs()
    low_close = (df["Low"] - df["Close"].shift()).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)

    atr = tr.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    plus_di = 100 * pd.Series(plus_dm, index=df.index).ewm(alpha=1 / period, adjust=False, min_periods=period).mean() / atr.replace(0, np.nan)
    minus_di = 100 * pd.Series(minus_dm, index=df.index).ewm(alpha=1 / period, adjust=False, min_periods=period).mean() / atr.replace(0, np.nan)

    dx = (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan) * 100
    adx = dx.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()

    df["PLUS_DI"] = plus_di
    df["MINUS_DI"] = minus_di
    df["ADX"] = adx
    return df


def classify_market_regime(df, trend_threshold=25, range_threshold=20):
    """
    只根據最新一筆 ADX 判斷市場狀態，回傳 ("trending"/"ranging"/"transitional", adx值)。
    這是分類，不是預測——用的是「已經發生」的趨勢強度，不代表未來會延續。
    """
    latest_adx = df["ADX"].iloc[-1] if "ADX" in df.columns else None
    if latest_adx is None or pd.isna(latest_adx):
        return "unknown", None
    if latest_adx >= trend_threshold:
        return "trending", float(latest_adx)
    elif latest_adx < range_threshold:
        return "ranging", float(latest_adx)
    else:
        return "transitional", float(latest_adx)

def calc_ewma_volatility(df, lambda_param=0.94):
    """採用 RiskMetrics 標準 EWMA 模型計算最新年化與日動態波動率"""
    log_returns = np.log(df["Close"] / df["Close"].shift(1)).dropna()
    variance = np.zeros_like(log_returns)
    variance[0] = log_returns.var()
    
    for t in range(1, len(log_returns)):
        variance[t] = lambda_param * variance[t-1] + (1 - lambda_param) * (log_returns.iloc[t-1] ** 2)
        
    latest_daily_vol = np.sqrt(variance[-1])
    return latest_daily_vol  # 回傳最新一日 EWMA 標準差

def technical_summary(df, regime="unknown", adx_value=None):
    """
    技術面評分依「市場狀態」動態調整邏輯：
    - trending（ADX>=25）：維持原設計，趨勢+MACD為主要權重，RSI/KD僅描述狀態不逆勢操作
      （理由：學術與實務上都建議「趨勢明確時不要用超買超賣去反著做」）。
    - ranging（ADX<20）：趨勢排列與MACD權重減半，改用RSI/KD的均值回歸(mean-reversion)訊號
      （理由：盤整市價格傾向在區間內來回，超買超賣更容易真的回歸）。
    - transitional/unknown：維持原設計，但不做加權，因為狀態本身不明確。
    """
    latest = df.iloc[-1]
    prev = df.iloc[-2] if len(df) >= 2 else latest
    notes = []
    score = 0

    is_ranging = regime == "ranging"
    trend_weight = 0.5 if is_ranging else 1.0
    # MACD 是純趨勢型指標，在真正的盤整市不是「打折」而是「容易連續假訊號」，
    # 因此盤整市直接停用其加減分（而非打對折），避免用銀行家捨入 round(0.5)=0
    # 這種容易誤導的寫法去偽裝「減半」。
    macd_enabled = not is_ranging

    if adx_value is not None:
        regime_label = {"trending": "趨勢市", "ranging": "盤整市", "transitional": "過渡帶"}.get(regime, "無法判斷")
        notes.append(f"市場狀態(ADX={adx_value:.1f}): {regime_label}"
                     + ("，技術面改採均值回歸邏輯" if is_ranging else ""))

    # 趨勢：盤整市權重減半（四捨五入取整數分數，避免出現0.5分混亂總分計算）
    if latest["Close"] > latest["MA5"] > latest["MA20"] > latest["MA60"]:
        notes.append("股價站上所有均線，呈多頭排列" + ("（盤整市，權重減半）" if is_ranging else ""))
        score += round(2 * trend_weight)
    elif latest["Close"] < latest["MA5"] < latest["MA20"] < latest["MA60"]:
        notes.append("股價跌破所有均線，呈空頭排列" + ("（盤整市，權重減半）" if is_ranging else ""))
        score -= round(2 * trend_weight)
    else:
        notes.append("均線排列混合，趨勢尚未形成一致方向")

    # RSI：趨勢市僅描述狀態；盤整市啟用均值回歸訊號（超買偏空、超賣偏多）
    rsi = latest["RSI"]
    if pd.notna(rsi):
        if rsi >= 70:
            if is_ranging:
                notes.append(f"RSI={rsi:.1f}，盤整市超買，均值回歸訊號偏空")
                score -= 1
            else:
                notes.append(f"RSI={rsi:.1f}，動能強但已進入過熱區；不直接視為看空訊號")
        elif rsi <= 30:
            if is_ranging:
                notes.append(f"RSI={rsi:.1f}，盤整市超賣，均值回歸訊號偏多")
                score += 1
            else:
                notes.append(f"RSI={rsi:.1f}，動能弱且進入超賣區；不直接視為買進訊號")
        elif rsi >= 55:
            notes.append(f"RSI={rsi:.1f}，動能偏強")
        elif rsi <= 45:
            notes.append(f"RSI={rsi:.1f}，動能偏弱")
        else:
            notes.append(f"RSI={rsi:.1f}，動能中性")

    # MACD：盤整市權重減半
    if latest["MACD"] > latest["MACD_signal"]:
        if latest["MACD_hist"] >= prev["MACD_hist"]:
            if macd_enabled:
                notes.append("MACD 位於信號線之上且柱體擴大，正向動能增強")
                score += 1
            else:
                notes.append("MACD 位於信號線之上且柱體擴大，但盤整市中趨勢型指標易生假訊號，不納入評分")
        else:
            notes.append("MACD 位於信號線之上，但正向動能正在收斂")
    else:
        if latest["MACD_hist"] <= prev["MACD_hist"]:
            if macd_enabled:
                notes.append("MACD 位於信號線之下且柱體走弱，負向動能增強")
                score -= 1
            else:
                notes.append("MACD 位於信號線之下且柱體走弱，但盤整市中趨勢型指標易生假訊號，不納入評分")
        else:
            notes.append("MACD 位於信號線之下，但負向動能正在收斂")

    # KD：盤整市下，黃金/死亡交叉本身就是均值回歸訊號的一種，維持原邏輯不額外加權
    # （因為KD交叉在盤整市原本就比趨勢市更常見、更容易失敗，這是KD這個指標本身的已知限制，
    #   不透過加權處理，而是提醒使用者盤整市的KD交叉訊號更需要小心）
    if all(pd.notna(x) for x in [latest["K"], latest["D"], prev["K"], prev["D"]]):
        golden_cross = prev["K"] <= prev["D"] and latest["K"] > latest["D"]
        death_cross = prev["K"] >= prev["D"] and latest["K"] < latest["D"]
        if golden_cross and latest["K"] < 80:
            notes.append("KD 近期黃金交叉且尚未過熱")
            score += 1
        elif death_cross and latest["K"] > 20:
            notes.append("KD 近期死亡交叉")
            score -= 1
        else:
            notes.append(f"KD 目前 K={latest['K']:.1f} / D={latest['D']:.1f}，無新增交叉訊號")

    # 成交量必須搭配價格方向判讀，不再單純「量增 = 加分」
    vol5 = df["Volume"].tail(5).mean()
    vol20 = df["Volume"].tail(20).mean()
    price_5d = latest["Close"] / df["Close"].iloc[-6] - 1 if len(df) >= 6 else 0
    if pd.notna(vol5) and pd.notna(vol20) and vol20 > 0 and vol5 > vol20 * 1.2:
        if price_5d > 0:
            notes.append("近期量增且近 5 日價格上漲，量價偏多")
            score += 1
        elif price_5d < 0:
            notes.append("近期量增但近 5 日價格下跌，賣壓偏重")
            score -= 1
        else:
            notes.append("近期成交量放大，但價格方向不明")

    return notes, score


# ----------------------------
# 基本面分析
# ----------------------------

def fundamental_summary(ticker_obj, current_price=None):
    """
    基本面評分刻意避免用固定 P/E 15/30 判斷所有產業。
    P/E 只作估值資訊；評分較著重獲利品質、成長一致性與財務風險。
    若要做嚴格估值判斷，應再加入同業與自身歷史估值分位數。

    回傳 (notes, score, info, metrics)：metrics 是給 Excel 匯出等下游用途的
    乾淨數值字典，不要從 notes 文字反推數字——文字措辭以後可能會改，
    字典欄位是穩定介面。
    """
    info = ticker_obj.info
    notes = []
    score = 0
    metrics = {}

    def g(key):
        return info.get(key, None)

    pe = g("trailingPE")
    forward_pe = g("forwardPE")
    roe = g("returnOnEquity")
    profit_margin = g("profitMargins")
    rev_growth = g("revenueGrowth")
    earnings_growth = g("earningsGrowth")
    div_yield = g("dividendYield")
    debt_to_equity = g("debtToEquity")
    sector = g("sector")
    industry = g("industry")

    if sector or industry:
        notes.append(f"產業: {sector or 'N/A'} / {industry or 'N/A'}")
    metrics["sector"] = sector
    metrics["industry"] = industry

    if pe is not None:
        notes.append(f"本益比 (P/E): {pe:.2f}（僅列示，不使用跨產業固定門檻評分）")

    if forward_pe is not None:
        notes.append(f"預估本益比 (Forward P/E): {forward_pe:.2f}")
        if pe is not None and pe > 0 and forward_pe > 0:
            diff = forward_pe / pe - 1
            if diff <= -0.10:
                notes.append("Forward P/E 明顯低於目前 P/E，市場預估獲利可能改善；仍需核對分析師預估可靠度")
            elif diff >= 0.10:
                notes.append("Forward P/E 高於目前 P/E，可能反映未來獲利預估轉弱；需進一步確認原因")

    # ROE 與利潤率需搭配，不讓單一高 ROE（可能來自高槓桿）主導分數
    if roe is not None:
        notes.append(f"股東權益報酬率 (ROE): {roe*100:.2f}%")
    if profit_margin is not None:
        notes.append(f"淨利率: {profit_margin*100:.2f}%")

    if roe is not None and profit_margin is not None:
        if roe >= 0.15 and profit_margin > 0:
            score += 1
            notes.append("ROE 與淨利率同為正向，獲利品質偏佳")
        elif roe < 0 or profit_margin < 0:
            score -= 1
            notes.append("ROE 或淨利率為負，獲利品質需留意")
    elif roe is not None:
        if roe >= 0.15:
            score += 1
        elif roe < 0:
            score -= 1

    if rev_growth is not None:
        notes.append(f"營收成長率 (YoY): {rev_growth*100:.2f}%")
    if earnings_growth is not None:
        notes.append(f"獲利成長率 (YoY): {earnings_growth*100:.2f}%")

    # 成長評分看營收與獲利是否同向，比單看營收 >10% 更穩健
    if rev_growth is not None and earnings_growth is not None:
        if rev_growth > 0 and earnings_growth > 0:
            score += 1
            notes.append("營收與獲利同步成長，成長品質偏正向")
        elif rev_growth < 0 and earnings_growth < 0:
            score -= 1
            notes.append("營收與獲利同步衰退，成長動能偏弱")
        elif rev_growth > 0 and earnings_growth < 0:
            score -= 1
            notes.append("營收成長但獲利衰退，可能有毛利率/費用壓力，需進一步拆解")
        elif rev_growth < 0 and earnings_growth > 0:
            notes.append("營收衰退但獲利改善，可能來自成本或一次性因素，需確認可持續性")
    elif rev_growth is not None:
        if rev_growth > 0:
            score += 1
        elif rev_growth < 0:
            score -= 1

    if div_yield is not None:
        raw_yield = float(div_yield)
        yield_pct = raw_yield * 100 if raw_yield <= 1 else raw_yield
        dividend_rate = g("dividendRate")
        cross_check_pct = None
        if dividend_rate is not None and current_price and current_price > 0:
            try:
                cross_check_pct = float(dividend_rate) / float(current_price) * 100
            except (TypeError, ValueError, ZeroDivisionError):
                cross_check_pct = None
        if cross_check_pct is not None:
            if abs(cross_check_pct - yield_pct) > max(1.0, abs(cross_check_pct) * 0.20):
                notes.append(f"殖利率修正：API原值過異，採股利/現價計算值={cross_check_pct:.2f}%")
                yield_pct = cross_check_pct
            else:
                notes.append(f"殖利率: {yield_pct:.2f}%")
        elif yield_pct > 20:
            notes.append(
                f"殖利率: API 顯示 {yield_pct:.2f}%，但缺少可交叉驗證的股利/現價資料；"
                "數值異常，暫不納入評分"
            )
        else:
            notes.append(f"殖利率: {yield_pct:.2f}%（僅列示，不直接作多空評分）")

        # 交叉驗證兩個欄位「互相同意」不代表兩者都對——如果 dividendRate 與
        # current_price 幣別基準不一致（常見於 .TW ADR 相關欄位），兩個算出來的
        # 百分比可能剛好都被同一個錯誤放大，導致上面的差異比對被騙過去。
        # 因此不論走哪個分支，最終數值仍要過一次絕對合理性檢查。
        # 負值同樣視為異常：正常殖利率不會是負的，出現負值代表欄位定義或
        # 單位換算有問題，跟「超過20%上限」同一等級處理，不納入評分。
        yield_abnormal = yield_pct > 20 or yield_pct < 0
        if yield_pct < 0:
            notes.append(f"⚠ 殖利率最終值 {yield_pct:.2f}% 為負，屬於資料異常（正常殖利率不會是負值）；"
                          "此數值不納入基本面評分，且不建議直接引用。")
        elif yield_abnormal:
            notes.append(
                f"⚠ 殖利率最終值 {yield_pct:.2f}% 超過合理上限（多數市場殖利率極少超過 15~20%），"
                "即使已通過與 dividendRate 的交叉比對，仍可能是幣別或欄位基準錯位所致；"
                "此數值不納入基本面評分，且不建議直接引用。"
            )
        metrics["dividend_yield_pct"] = yield_pct
        metrics["dividend_yield_abnormal"] = yield_abnormal

    if debt_to_equity is not None:
        try:
            de = float(debt_to_equity)
            notes.append(f"負債權益比: {de:.2f}%（yfinance 常見定義：百分比數值，非 16.5 倍）")
            if de < 0:
                notes.append("⚠ 負債權益比為負，可能代表股東權益為負(資不抵債)或資料定義異常；"
                              "此數值不納入評分，且是比「偏高」更需要優先確認的財務風險訊號")
            elif de > 500:
                notes.append("⚠ 負債權益比極端異常，可能是資料定義/單位問題；不要直接拿此數字判斷公司槓桿")
            elif de > 250:
                score -= 1
                notes.append("負債權益比偏高；應搭配利息保障倍數、自由現金流與同業比較")
            metrics["debt_to_equity"] = de
            metrics["debt_to_equity_abnormal"] = de < 0 or de > 500
        except (TypeError, ValueError):
            notes.append("負債權益比資料格式無法解析，暫不納入評分")

    notes.append("資料基準提醒: 本工具之 P/E、ROE、成長率等多數比率取自 yfinance 提供之 TTM(近四季)或最新一季相關欄位，"
                 "實際計算基準請以財報原文為準，跨公司比較時請確認基準期間一致")

    notes.append("估值提醒: P/E 與 Forward P/E 是不同期間的獲利基礎；不能只因 Forward P/E 較低就判定股價便宜。"
                 "若要做估值，下一版應加入同業與自身歷史 P/E 分位數。")

    metrics.update({
        "pe": pe, "forward_pe": forward_pe,
        "roe_pct": roe * 100 if roe is not None else None,
        "profit_margin_pct": profit_margin * 100 if profit_margin is not None else None,
        "rev_growth_pct": rev_growth * 100 if rev_growth is not None else None,
        "earnings_growth_pct": earnings_growth * 100 if earnings_growth is not None else None,
    })
    return notes, score, info, metrics


# ----------------------------
# 籌碼面分析
# ----------------------------

def chip_summary_us(ticker_obj, info=None):
    """
    美股籌碼面: 法人持股、內部人買賣、主要股東。

    info 可由呼叫端傳入(通常是 fundamental_summary 已經抓過的同一份 info)，
    避免同一次分析對同一檔股票重複呼叫 yfinance 的 .info（這個屬性慢、
    偶爾 timeout、且欄位不一定齊全，能少抓一次就少一次失敗風險）。
    若沒有傳入，才自己抓一次。

    insider_transactions 的 Text 欄位用關鍵字判斷買賣方向，這只是粗略分類，
    不是嚴謹的交易類型解析：無法區分薪酬型/期權型/10b5-1 自動賣出這類
    「賣出但不代表看空」的情況，也可能有無法辨識的措辭，一律歸類「未分類」，
    不強行套進買/賣其中一類。
    """
    notes = []
    score = 0
    metrics = {}

    if info is None:
        try:
            info = ticker_obj.info
        except Exception:
            info = {}

    try:
        inst_pct = info.get("heldPercentInstitutions") if info else None
        insider_pct = info.get("heldPercentInsiders") if info else None

        if inst_pct is not None:
            notes.append(f"法人持股比例: {inst_pct*100:.2f}%（持股高低本身不直接視為多空訊號）")
            metrics["institution_holding_pct"] = inst_pct * 100
        if insider_pct is not None:
            notes.append(f"內部人持股比例: {insider_pct*100:.2f}%")
            metrics["insider_holding_pct"] = insider_pct * 100
    except Exception:
        pass

    buy_keywords = ("buy", "purchase", "purchased", "acquired")
    sell_keywords = ("sale", "sell", "sold", "disposed", "disposition")

    try:
        insider_tx = ticker_obj.insider_transactions
        if insider_tx is not None and not insider_tx.empty:
            recent = insider_tx.head(10)
            buy_count = 0
            sell_count = 0
            unknown_count = 0
            for _, row in recent.iterrows():
                text = str(row.get("Text", "")).strip().lower()
                if any(k in text for k in sell_keywords):
                    sell_count += 1
                elif any(k in text for k in buy_keywords):
                    buy_count += 1
                else:
                    unknown_count += 1
            notes.append(f"近期內部人交易: 買進 {buy_count} 筆 / 賣出 {sell_count} 筆 / 未分類 {unknown_count} 筆"
                         f"（近 {len(recent)} 筆，僅依關鍵字粗略判斷方向，不分辨薪酬/期權/10b5-1等交易性質）")
            metrics["insider_buy_count"] = buy_count
            metrics["insider_sell_count"] = sell_count
            metrics["insider_unknown_count"] = unknown_count
            if buy_count > 0 and buy_count > sell_count:
                score += 1
                notes.append("近期內部人主動買進較多，視為弱正向訊號")
            elif sell_count > buy_count:
                notes.append("近期內部人賣出較多，但賣出可能來自薪酬、稅務、10b5-1 計畫或資產配置，不直接扣分")
    except Exception:
        pass

    if not notes:
        notes.append("無法取得法人/內部人籌碼資料")

    return notes, score, None, metrics


def _safe_int(value):
    if value is None:
        return 0
    text = str(value).replace(",", "").replace("--", "0").strip()
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return 0


_CHIP_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chip_cache")


def _append_chip_cache(stock_code, records):
    """
    把這次抓到的每日三大法人數字寫入本地 CSV（去重，同一天不重複寫入）。
    這個快取會隨著你每天執行程式逐漸累積歷史——不是一次到位。
    第一次執行時，斜率類特徵幾乎一定是 None，這是預期行為，不是 bug。
    """
    try:
        os.makedirs(_CHIP_CACHE_DIR, exist_ok=True)
        path = os.path.join(_CHIP_CACHE_DIR, f"{stock_code}.csv")
        existing = set()
        if os.path.exists(path):
            existing_df = pd.read_csv(path, dtype=str)
            existing = set(existing_df["date"].tolist())
        new_rows = [r for r in records if r[0] not in existing]
        if new_rows:
            df_new = pd.DataFrame(new_rows, columns=["date", "foreign_net", "trust_net", "dealer_net", "total_net"])
            write_header = not os.path.exists(path)
            df_new.to_csv(path, mode="a", header=write_header, index=False)
    except Exception:
        pass  # 快取寫入失敗不應該讓主流程中斷，籌碼分析仍要能正常輸出


def _foreign_slope_from_cache(stock_code, lookback=5):
    """
    讀取本地快取，計算外資買賣超的線性回歸斜率（單位：股/交易日）。
    累積筆數不足 lookback 時回傳 None，並附上目前累積筆數，讓使用者知道
    還需要再執行幾次程式才能開始看到這個特徵。
    """
    path = os.path.join(_CHIP_CACHE_DIR, f"{stock_code}.csv")
    if not os.path.exists(path):
        return None, 0
    try:
        df = pd.read_csv(path)
        df = df.sort_values("date").tail(lookback)
        if len(df) < lookback:
            return None, len(df)
        x = np.arange(len(df))
        y = df["foreign_net"].astype(float).values
        slope = float(np.polyfit(x, y, 1)[0])
        return slope, len(df)
    except Exception:
        return None, 0


def _append_concentration_cache(stock_code, date_str, ratio_pct):
    """
    籌碼集中度(外資買賣超佔當日成交量比重)的獨立快取檔案。

    刻意不併入既有的 {stock_code}.csv（外資/投信/自營淨額快取），而是
    另開 {stock_code}_concentration.csv：因為比重需要當日成交量，這個
    資訊只有在 analyze() 裡同時拿到 hist 才能算出來，跟 chip_summary_tw()
    自己就能算的三大法人淨額不是同一個資料來源、同一個時間點寫入；
    分開存也避免更動既有 {stock_code}.csv 的欄位結構、影響到已經在用的
    foreign_5d_slope 邏輯。

    這是「報告層級指標」，不是模型訓練特徵——TWSE 免費端點只給最新快照，
    無法回填完整歷史，天花板跟 foreign_5d_slope 一樣，只能日復一日累積。
    """
    try:
        os.makedirs(_CHIP_CACHE_DIR, exist_ok=True)
        path = os.path.join(_CHIP_CACHE_DIR, f"{stock_code}_concentration.csv")
        existing = set()
        if os.path.exists(path):
            existing_df = pd.read_csv(path, dtype=str)
            existing = set(existing_df["date"].tolist())
        if date_str not in existing:
            df_new = pd.DataFrame([[date_str, ratio_pct]], columns=["date", "ratio_pct"])
            write_header = not os.path.exists(path)
            df_new.to_csv(path, mode="a", header=write_header, index=False)
    except Exception:
        pass  # 快取寫入失敗不應該讓主流程中斷


def _concentration_slope_from_cache(stock_code, lookback=5):
    """讀取籌碼集中度本地快取，計算近 lookback 個交易日的線性回歸斜率(個百分點/交易日)。"""
    path = os.path.join(_CHIP_CACHE_DIR, f"{stock_code}_concentration.csv")
    if not os.path.exists(path):
        return None, 0
    try:
        df = pd.read_csv(path)
        df = df.sort_values("date").tail(lookback)
        if len(df) < lookback:
            return None, len(df)
        x = np.arange(len(df))
        y = df["ratio_pct"].astype(float).values
        slope = float(np.polyfit(x, y, 1)[0])
        return slope, len(df)
    except Exception:
        return None, 0


def chip_summary_tw(stock_code, trading_days=5):
    """上市股票：抓取最近數個交易日 TWSE T86，避免用單一交易日判斷籌碼。"""
    notes = []
    score = 0
    if requests is None:
        return ["需安裝 requests 套件才能取得台股籌碼資料 (pip install requests)"], 0, None, {}

    code = stock_code.upper().replace(".TW", "")
    records = []

    # 往前找最多 14 個曆日，收集最近 trading_days 個有效交易日
    for i in range(14):
        if len(records) >= trading_days:
            break
        date = datetime.date.today() - datetime.timedelta(days=i)
        date_str = date.strftime("%Y%m%d")
        url = f"https://www.twse.com.tw/rwd/zh/fund/T86?response=json&date={date_str}&selectType=ALL"
        try:
            resp = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
            data = resp.json()
        except Exception:
            continue
        finally:
            time.sleep(0.3)  # 避免連續請求被 TWSE 視為異常流量
        if data.get("stat") != "OK":
            continue
        for row in data.get("data", []):
            if row and row[0].strip() == code:
                try:
                    foreign_net = _safe_int(row[4])
                    trust_net = _safe_int(row[10])
                    dealer_net = _safe_int(row[11])
                    total = foreign_net + trust_net + dealer_net
                    records.append((date_str, foreign_net, trust_net, dealer_net, total))
                except IndexError:
                    pass
                break

    if not records:
        return ["查無近期 TWSE 三大法人資料（可能為非交易日、代碼錯誤或資料暫時不可用）"], 0, None, {}

    latest = records[0]
    notes.append(f"最新資料日期: {latest[0]}")
    notes.append(f"最新日 外資/投信/自營商: {latest[1]:,} / {latest[2]:,} / {latest[3]:,} 股")

    total_foreign = sum(r[1] for r in records)
    total_trust = sum(r[2] for r in records)
    total_dealer = sum(r[3] for r in records)
    total_all = sum(r[4] for r in records)
    buy_days = sum(1 for r in records if r[4] > 0)
    sell_days = sum(1 for r in records if r[4] < 0)

    notes.append(f"近 {len(records)} 個交易日外資累計: {total_foreign:,} 股")
    notes.append(f"近 {len(records)} 個交易日投信累計: {total_trust:,} 股")
    notes.append(f"近 {len(records)} 個交易日自營商累計: {total_dealer:,} 股")
    notes.append(f"近 {len(records)} 個交易日三大法人合計: {total_all:,} 股（買超 {buy_days} 日 / 賣超 {sell_days} 日）")

    # 多日累計 + 持續性，避免單日雜訊
    if total_all > 0 and buy_days >= max(3, len(records) - 2):
        score += 1
        notes.append("法人多日累計與持續性偏買方")
    elif total_all < 0 and sell_days >= max(3, len(records) - 2):
        score -= 1
        notes.append("法人多日累計與持續性偏賣方")
    else:
        notes.append("法人多日方向分歧，暫不加減分")

    # 寫入本地快取，供 Foreign_5D_Slope 這類需要「歷史時間序列」的特徵使用。
    # 這個快取會隨著每天執行逐步累積，不是一次到位。
    _append_chip_cache(code, records)
    slope, cache_n = _foreign_slope_from_cache(code, lookback=5)
    if slope is not None:
        direction = "轉強" if slope > 0 else "轉弱" if slope < 0 else "持平"
        notes.append(f"外資買賣超 5 日斜率(本地快取): {slope:,.0f} 股/交易日（{direction}）")
    else:
        notes.append(f"外資買賣超斜率: 本地快取僅累積 {cache_n} 個交易日，需累積滿 5 個交易日"
                     "（即連續執行本程式 5 個不同交易日）才會開始計算，這是預期行為")

    # 三大法人合計買賣超「連續同方向天數」：從最新一筆(records[0])往回數，
    # 方向(買超/賣超)持續一致的天數。只在這次抓到的 records 窗格內計算，
    # 不是無限回溯的歷史streak，窗格外是否延續無法得知，這裡如實反映限制。
    streak = 0
    streak_label = "持平"
    if records:
        first_net = records[0][4]
        first_is_buy = first_net > 0
        first_is_sell = first_net < 0
        for r in records:
            if first_is_buy and r[4] > 0:
                streak += 1
            elif first_is_sell and r[4] < 0:
                streak += 1
            else:
                break
        streak_label = "買超" if first_is_buy else "賣超" if first_is_sell else "持平"
        notes.append(f"三大法人合計連續{streak_label} {streak} 個交易日（僅限本次抓取的近 {len(records)} 日窗格內）")

    metrics = {
        "foreign_net_latest": latest[1],
        "trust_net_latest": latest[2],
        "dealer_net_latest": latest[3],
        "foreign_5d_slope": slope,
        "net_streak_days": streak,
        "net_streak_direction": streak_label,
    }
    return notes, score, latest[0], metrics


def chip_summary_two(stock_code):
    """上櫃股票：使用櫃買中心 OpenAPI，不再誤套美股法人資料。"""
    notes = []
    score = 0
    if requests is None:
        return ["需安裝 requests 套件才能取得上櫃籌碼資料 (pip install requests)"], 0, None, {}

    code = stock_code.upper().replace(".TWO", "")
    url = "https://www.tpex.org.tw/openapi/v1/tpex_3insti_daily_trading"
    try:
        resp = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        rows = resp.json()
    except Exception as e:
        return [f"櫃買中心三大法人 OpenAPI 暫時無法取得: {type(e).__name__}"], 0, None, {}

    target = None
    for row in rows if isinstance(rows, list) else []:
        # OpenAPI 欄位可能隨版本以中文名稱回傳，因此同時檢查常見代碼欄位
        values = {str(k): v for k, v in row.items()}
        code_value = (values.get("SecuritiesCompanyCode") or values.get("證券代號") or
                      values.get("Code") or values.get("股票代號"))
        if str(code_value).strip() == code:
            target = values
            break

    if not target:
        return ["查無近期 TPEx 三大法人資料；已正確辨識為上櫃股票，不會再誤用美股籌碼資料"], 0, None, {}

    def pick(*keys):
        for k in keys:
            if k in target:
                return target[k]
        return None

    date_val = pick("Date", "日期", "資料日期")
    foreign = _safe_int(pick("Foreign_Investor_Buy_Sell", "外資及陸資買賣超股數", "外資買賣超股數"))
    trust = _safe_int(pick("Investment_Trust_Buy_Sell", "投信買賣超股數"))
    dealer = _safe_int(pick("Dealer_Buy_Sell", "自營商買賣超股數"))

    # 如果欄位名稱與預期不同，至少完整列出可辨識的三大法人合計欄位
    total = _safe_int(pick("Total", "三大法人買賣超股數", "三大法人買賣超"))
    if total == 0 and any([foreign, trust, dealer]):
        total = foreign + trust + dealer

    if date_val:
        notes.append(f"資料日期: {date_val}")
    if any([foreign, trust, dealer]):
        notes.append(f"外資/投信/自營商: {foreign:,} / {trust:,} / {dealer:,} 股")
    notes.append(f"三大法人合計買賣超: {total:,} 股")
    notes.append("上櫃 OpenAPI 目前此模組以最新日資料為主；不與美股持股資料混用")

    # 單日資料只做弱訊號，避免誤判
    if total > 0:
        notes.append("最新日法人偏買方（單日弱訊號，不加分）")
    elif total < 0:
        notes.append("最新日法人偏賣方（單日弱訊號，不扣分）")

    metrics = {"foreign_net_latest": foreign, "trust_net_latest": trust, "dealer_net_latest": dealer}
    return notes, score, str(date_val) if date_val else None, metrics


def chip_summary(ticker_symbol, ticker_obj, info=None):
    symbol = ticker_symbol.upper()
    if symbol.endswith(".TW"):
        return chip_summary_tw(symbol)
    if symbol.endswith(".TWO"):
        return chip_summary_two(symbol)
    return chip_summary_us(ticker_obj, info=info)


def _parse_chip_date(date_str):
    """將籌碼面回傳的日期字串（如 20260818）轉為 date 物件，格式不明則回傳 None。"""
    if not date_str:
        return None
    text = str(date_str).strip()
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


# 台股休市日。只列出已確認的日期；清單過期時 next_trading_day_estimate()
# 會明確提示，而不是默默算出錯誤日期。
# 維護來源：TWSE 每年公告的「有價證券集中交易市場開（休）市日期表」。
TW_HOLIDAYS = {
    # 2026 年（請每年更新；未列出的日期一律視為有開市）
    datetime.date(2026, 1, 1),                                        # 元旦
    datetime.date(2026, 2, 14), datetime.date(2026, 2, 16),
    datetime.date(2026, 2, 17), datetime.date(2026, 2, 18),
    datetime.date(2026, 2, 19), datetime.date(2026, 2, 20),           # 春節
    datetime.date(2026, 2, 27), datetime.date(2026, 2, 28),           # 和平紀念日
    datetime.date(2026, 4, 3), datetime.date(2026, 4, 6),             # 清明/兒童節
    datetime.date(2026, 5, 1),                                        # 勞動節
    datetime.date(2026, 6, 19),                                       # 端午節
    datetime.date(2026, 9, 25),                                       # 中秋節
    datetime.date(2026, 10, 9), datetime.date(2026, 10, 26),          # 國慶/光復節
}

TW_HOLIDAY_COVERAGE_UNTIL = datetime.date(2026, 12, 31)


def next_trading_day_estimate(basis_date, holidays=None):
    """
    從 basis_date（資料基準日，即最新收盤價所屬的那一天）往後推到下一個
    交易日，跳過週六日「與」國定假日。

    v3.7 修正：v3.6 只跳週末，遇到連假會把「預測目標日」標錯。那不是
    顯示問題——事後用 tools/log_review.py 回填實際結果時會對錯日子，
    是會靜靜汙染整份驗證統計的那種錯。

    仍然保留「這只是估計」的定位：假日清單維護到期之後，會在回傳的
    提醒訊息裡明講，讓你自己核對，而不是假裝這個日期絕對正確。

    回傳 (日期, 提醒訊息)。沒有提醒時訊息為空字串。
    """
    holidays = TW_HOLIDAYS if holidays is None else holidays
    d = basis_date + datetime.timedelta(days=1)
    while d.weekday() >= 5 or d in holidays:  # 5=Sat, 6=Sun
        d += datetime.timedelta(days=1)

    note = ""
    if d > TW_HOLIDAY_COVERAGE_UNTIL:
        note = (f"假日清單只維護到 {TW_HOLIDAY_COVERAGE_UNTIL}，"
                f"{d} 已超出涵蓋範圍，請自行核對是否為交易日")
    return d, note


# ----------------------------
# 隔日歷史波動參考區間(非預測值)
# ----------------------------

def _empirical_interval(returns):
    """用歷史報酬分位數建立經驗式區間，不強迫套用常態分布。"""
    r = pd.Series(returns).dropna()
    if len(r) < 30:
        return None
    q16, q84 = r.quantile([0.16, 0.84])
    q025, q975 = r.quantile([0.025, 0.975])
    q50 = r.quantile(0.5)
    return {"low68": float(q16), "high68": float(q84),
            "low95": float(q025), "high95": float(q975),
            "median": float(q50),
            "mean": float(r.mean()), "std": float(r.std()), "n": int(len(r))}


def next_day_range(df, window=60):
    """用前 window 日 Close-to-Close log return 的分位數建立歷史經驗區間。"""
    df = calc_atr(df.copy())
    latest_close = float(df["Close"].iloc[-1])
    latest_atr = float(df["ATR"].iloc[-1])
    log_returns = np.log(df["Close"] / df["Close"].shift(1)).dropna()
    interval = _empirical_interval(log_returns.tail(window))
    if interval is None:
        return None
    return {
        "close": latest_close, "atr": latest_atr,
        "daily_std_pct": interval["std"] * 100,
        "mean_return_pct": interval["mean"] * 100,
        "median_price": latest_close * np.exp(interval["median"]),
        "range_68": (latest_close * np.exp(interval["low68"]), latest_close * np.exp(interval["high68"])),
        "range_95": (latest_close * np.exp(interval["low95"]), latest_close * np.exp(interval["high95"])),
        "window": interval["n"]
    }


def backtest_confidence(df, lookback=500, window=60):
    """嚴格只用每個交易日前的 window 日資料建立經驗區間，再測下一日覆蓋率。"""
    close = df["Close"].astype(float)
    log_returns = np.log(close / close.shift(1))
    hit_68 = hit_95 = total = 0
    n = len(df)
    start_idx = max(window + 1, n - lookback)
    for t in range(start_idx, n):
        past_returns = log_returns.iloc[t-window:t].dropna()
        interval = _empirical_interval(past_returns)
        if interval is None:
            continue
        actual = float(log_returns.iloc[t])
        hit_68 += int(interval["low68"] <= actual <= interval["high68"])
        hit_95 += int(interval["low95"] <= actual <= interval["high95"])
        total += 1
    if total == 0:
        return None
    h68 = hit_68 / total * 100
    h95 = hit_95 / total * 100
    return {"sample_size": total, "empirical_conf_68": h68, "empirical_conf_95": h95,
            "target_68_gap": h68 - 68.0, "target_95_gap": h95 - 95.0}


# ----------------------------
# 隔日方向機率預估（樣本外回測，非保證預測）
# ----------------------------
#
# 設計原則（務必保留這段說明，這是本模組能不能被信任的關鍵）：
#   1. 只用「當天收盤時已知」的資訊做特徵，不用任何未來資料 → 避免 look-ahead bias。
#   2. 用 walk-forward（走動式）方式回測：每一次預測，模型都只用「這一天之前」
#      的資料訓練，完全模擬「如果你當時真的照這個模型做」的情境。
#   3. 一定要跟 naive baseline（例如永遠猜漲）比較。因為多數股票長期上漲天數
#      本來就略多於下跌天數，如果模型準確率沒有明顯超過這個 baseline，
#      代表模型沒有提供真正的資訊量，只是在利用市場本身的正偏態。
#   4. 這是「機率估計」，不是「保證預測」。即使樣本外準確率有 55%，
#      也代表有 45% 的機率判斷錯誤方向，不能當作下單訊號單獨使用。

def _logistic_regression_fit(X, y, n_iter=300, lr=0.3, l2=1e-3):
    """簡單的手刻邏輯迴歸（梯度下降 + L2 正則化），避免額外依賴 sklearn。"""
    n, d = X.shape
    w = np.zeros(d)
    b = 0.0
    for _ in range(n_iter):
        z = X @ w + b
        p = 1 / (1 + np.exp(-np.clip(z, -30, 30)))
        grad_w = X.T @ (p - y) / n + l2 * w
        grad_b = np.mean(p - y)
        w -= lr * grad_w
        b -= lr * grad_b
    return w, b


_INTERMARKET_TICKERS = {"SOX": "^SOX", "USDTWD": "TWD=X", "UST10Y": "^TNX"}
_INTERMARKET_CACHE = {}


def get_intermarket_features(index, period="3y"):
    """
    抓取大盤關聯指標(費城半導體指數/美元兌台幣/美國10年期公債殖利率)，
    計算相對20日均線的乖離率與5日變動率，reindex 對齊到傳入的 index。

    設計重點：
    - 同一個 period 只抓一次、存進模組層級快取，批次查詢多檔股票時共用，
      不會每檔股票都重打一次外部 API。
    - reindex(method="ffill") 只會用「目標日期(含)以前」最後一筆已知收盤價
      做對齊，不使用未來資料，避免 look-ahead bias。
    - 任一指標抓取失敗（例如網路問題），只會少那一組欄位，不會讓整個
      特徵工程或呼叫端失敗——回傳的 cols 清單只包含實際成功抓到的欄位。
    """
    if period not in _INTERMARKET_CACHE:
        raw = {}
        for key, sym in _INTERMARKET_TICKERS.items():
            try:
                h = yf.Ticker(sym).history(period=period, auto_adjust=True)
                if not h.empty:
                    raw[key] = h["Close"].sort_index()
            except Exception:
                continue
        _INTERMARKET_CACHE[period] = raw
    raw = _INTERMARKET_CACHE[period]

    feat = pd.DataFrame(index=index)
    cols = []
    for key, series in raw.items():
        try:
            aligned = series.reindex(index, method="ffill")
            ma20 = aligned.rolling(20).mean()
            dev_col = f"{key}_dev20"
            chg_col = f"{key}_chg5"
            feat[dev_col] = (aligned - ma20) / ma20
            feat[chg_col] = aligned.pct_change(5)
            cols += [dev_col, chg_col]
        except Exception:
            continue
    return feat, cols


def _build_direction_features(df, intermarket_feat=None):
    """
    建立預測「隔日方向」用的特徵，全部只用當天(含)以前的資料，不使用未來資訊。

    這一版改用「變化率/相對值」而非技術指標的絕對水準（例如用 rsi_diff
    而非 rsi 本身），理由：RSI/KD 這類有界指標(0~100)本身不是定態
    (stationary)時間序列，不同期間的「RSI=60」代表的市場意義並不穩定；
    用一階差分（變化的方向與速度）通常比用水準值更適合丟進線性/近線性模型。

    overnight_gap 特別針對台股：0050、2330 等權值股與美股 ADR 連動度高，
    隔夜跳空往往反映的是「隔夜已發生但台股尚未反應」的資訊，比純技術指標
    更貼近「隔日方向」這個預測目標的因果機制。

    intermarket_feat: 選用，get_intermarket_features() 回傳的特徵表，
    直接 join 進來（欄位名稱不會與這裡的既有欄位衝突）。
    """
    feat = pd.DataFrame(index=df.index)

    prev_close = df["Close"].shift(1)
    feat["overnight_gap"] = (df["Open"] - prev_close) / prev_close
    feat["intraday_return"] = (df["Close"] - df["Open"]) / df["Open"]

    log_ret = np.log(df["Close"] / df["Close"].shift(1))
    feat["volatility_5d"] = log_ret.rolling(5).std()
    feat["volatility_20d"] = log_ret.rolling(20).std()

    feat["rsi_diff"] = df["RSI"].diff(1)
    feat["macd_hist_slope"] = df["MACD_hist"].diff(1)

    vol20 = df["Volume"].rolling(20).mean()
    feat["volume_ratio"] = df["Volume"] / vol20

    feat["mom5"] = df["Close"].pct_change(5)

    # 市場狀態(依 ADX 判斷)當作特徵，讓模型知道現在是趨勢市還是盤整市，
    # 而不是對所有市場狀態套用同一組係數。邏輯與 classify_market_regime()
    # 一致，只是這裡套用到整段歷史(逐列判斷)，不只是最新一列。
    if "ADX" in df.columns:
        adx = df["ADX"]
    else:
        adx = calc_adx(df.copy())["ADX"]
    feat["regime_trending"] = (adx >= 25).astype(float)
    feat["regime_ranging"] = (adx < 20).astype(float)

    if intermarket_feat is not None:
        feat = feat.join(intermarket_feat)

    return feat


DIRECTION_FEATURE_COLS = [
    "overnight_gap", "intraday_return", "volatility_5d", "volatility_20d",
    "rsi_diff", "macd_hist_slope", "volume_ratio", "mom5",
    "regime_trending", "regime_ranging",
]


try:
    from sklearn.ensemble import RandomForestClassifier
    _SKLEARN_AVAILABLE = True
except ImportError:
    _SKLEARN_AVAILABLE = False


def _fit_predict_logistic(X_train, y_train, X_test):
    mu = X_train.mean(axis=0)
    sigma = X_train.std(axis=0)
    sigma[sigma == 0] = 1.0
    w, b = _logistic_regression_fit((X_train - mu) / sigma, y_train)
    x_t = (X_test - mu) / sigma
    return 1 / (1 + np.exp(-np.clip(x_t @ w + b, -30, 30)))


def _fit_predict_random_forest(X_train, y_train, X_test):
    """
    Random Forest 版本。刻意用淺樹 + 限制葉節點樣本數來壓低 overfitting 風險
    ——在只有幾百筆樣本、個位數特徵的資料量級下，樹模型比邏輯迴歸容易overfit，
    這裡的參數設定是保守方向，不是為了追求訓練集準確率好看。
    """
    model = RandomForestClassifier(
        n_estimators=200, max_depth=4, min_samples_leaf=20,
        random_state=42, n_jobs=-1,
    )
    model.fit(X_train, y_train)
    return model.predict_proba(X_test)[:, 1]


def _fit_final_and_get_importance(model, X_train, y_train, X_test):
    """
    只給 walk-forward 函式「預測明天」那一次最終訓練使用（每次呼叫只跑一次，
    不影響 walk-forward 回測迴圈本身的效能），額外把特徵重要性/係數一併算出來。

    - random_forest: 直接用 sklearn 的 feature_importances_（Gini重要性）。
    - logistic: 沒有 feature_importances_ 這種東西，用標準化後係數的絕對值
      做類比——數值越大代表該特徵對預測結果的線性影響力越大，但這不是
      嚴謹意義下的「重要性」（沒有考慮特徵間交互作用、共線性），僅供
      排序參考，不要拿來做因果解讀。

    回傳 (p_test, importances)，importances 為 numpy array，跟傳入的
    X_train 欄位順序一一對應。
    """
    if model == "random_forest":
        rf = RandomForestClassifier(
            n_estimators=200, max_depth=4, min_samples_leaf=20,
            random_state=42, n_jobs=-1,
        )
        rf.fit(X_train, y_train)
        p_test = rf.predict_proba(X_test)[:, 1]
        return p_test, rf.feature_importances_
    else:
        mu = X_train.mean(axis=0)
        sigma = X_train.std(axis=0)
        sigma[sigma == 0] = 1.0
        w, b = _logistic_regression_fit((X_train - mu) / sigma, y_train)
        x_t = (X_test - mu) / sigma
        p_test = 1 / (1 + np.exp(-np.clip(x_t @ w + b, -30, 30)))
        return p_test, np.abs(w)


def direction_probability_walkforward(df, train_window=250, retrain_every=10, model="logistic",
                                       intermarket_feat=None, intermarket_cols=None,
                                       drift_window=30, drift_threshold=0.5):
    """
    估計「明日上漲機率」，並用 walk-forward 回測誠實回報樣本外準確率。
    model="logistic"（預設，零額外依賴）或 model="random_forest"（需要 scikit-learn，
    若未安裝會回傳 None 並在 notes 說明，而不是直接報錯中斷整支程式）。
    回傳 None 代表資料量不足，或指定的模型無法使用。

    v3.6 新增（皆為在既有 walk-forward 回測基礎上「多算幾個統計量」，
    不改變原本的訓練/預測邏輯本身）：
    - 期望值(EV)：假設每次都照模型判斷方向交易，紀錄每次的隔日報酬，
      算出勝率、平均獲利、平均虧損、EV = 勝率*平均獲利 - 敗率*平均虧損。
      這是回測期間「如果照做」的粗略估計，不含手續費/税/滑價。
    - 近 drift_window 個交易日的滾動樣本外準確率，低於 drift_threshold
      視為模型可能失效，回傳 drift_alert=True，交由呼叫端決定要不要示警。
      這個數字來自「這次重新執行」的 walk-forward 回測最後幾筆，
      不是從外部歷史紀錄檔反推（避免紀錄檔缺筆/對齊問題）。
    """
    if model == "random_forest" and not _SKLEARN_AVAILABLE:
        return None

    df = calc_rsi(df.copy())
    df = calc_macd(df)
    feat = _build_direction_features(df, intermarket_feat=intermarket_feat)
    feature_cols = DIRECTION_FEATURE_COLS + (intermarket_cols or [])
    feat_valid = feat.dropna(subset=feature_cols)

    target = (df["Close"].shift(-1) > df["Close"]).astype(int)
    next_return = df["Close"].pct_change().shift(-1)  # 隔日簡單報酬(%)，用於EV估算，不是log報酬
    bt_data = feat_valid.join(target.rename("target")).join(next_return.rename("next_return")).dropna()

    if len(bt_data) < train_window + 30:
        return None

    X_all = bt_data[feature_cols].values
    y_all = bt_data["target"].values
    ret_all = bt_data["next_return"].values
    n = len(bt_data)

    fit_predict = _fit_predict_random_forest if model == "random_forest" else _fit_predict_logistic

    preds, actuals, trade_returns = [], [], []
    p_up_cache = None
    for t in range(train_window, n):
        if p_up_cache is None or (t - train_window) % retrain_every == 0:
            X_train = X_all[t - train_window:t]
            y_train = y_all[t - train_window:t]
            # 一次對整個 retrain 區間做預測，避免每天重複訓練造成的高額運算成本
            end = min(t + retrain_every, n)
            p_up_cache = fit_predict(X_train, y_train, X_all[t:end])
            cache_start = t
        p_up = p_up_cache[t - cache_start]
        pred = int(p_up >= 0.5)
        preds.append(pred)
        actuals.append(int(y_all[t]))
        # 若照這次模型判斷的方向交易，隔日實際報酬會是多少(正=獲利，負=虧損)
        direction_sign = 1 if pred == 1 else -1
        trade_returns.append(direction_sign * float(ret_all[t]))

    preds = np.array(preds)
    actuals = np.array(actuals)
    trade_returns = np.array(trade_returns)
    accuracy = float((preds == actuals).mean())
    always_up_acc = float((actuals == 1).mean())
    naive_baseline = max(always_up_acc, 1 - always_up_acc)

    win_mask = trade_returns > 0
    loss_mask = trade_returns < 0
    win_rate = float(win_mask.mean())
    avg_gain_pct = float(trade_returns[win_mask].mean() * 100) if win_mask.any() else None
    avg_loss_pct = float(-trade_returns[loss_mask].mean() * 100) if loss_mask.any() else None
    if avg_gain_pct is not None and avg_loss_pct is not None:
        ev_pct = win_rate * avg_gain_pct - (1 - win_rate) * avg_loss_pct
    else:
        ev_pct = None

    if len(preds) >= drift_window:
        recent_acc = float((preds[-drift_window:] == actuals[-drift_window:]).mean())
        drift_alert = recent_acc < drift_threshold
    else:
        recent_acc = None
        drift_alert = False

    # 用最新一筆「答案未知」的特徵（今天）重新訓練、預測明天，
    # 同時取得特徵重要性/係數（只在這裡多算一次，不影響前面 walk-forward 迴圈效能）。
    X_train_final = X_all[-train_window:]
    y_train_final = y_all[-train_window:]
    latest_feat = feat_valid[feature_cols].iloc[-1].values.reshape(1, -1)
    p_final, importances = _fit_final_and_get_importance(model, X_train_final, y_train_final, latest_feat)
    p_up_tomorrow = float(p_final[0])
    feature_importance = dict(zip(feature_cols, [float(v) for v in importances]))

    return {
        "model": model,
        "p_up_tomorrow": p_up_tomorrow,
        "walk_forward_accuracy": accuracy * 100,
        "naive_baseline_accuracy": naive_baseline * 100,
        "edge_over_baseline": (accuracy - naive_baseline) * 100,
        "sample_size": len(preds),
        "win_rate_pct": win_rate * 100,
        "avg_gain_pct": avg_gain_pct,
        "avg_loss_pct": avg_loss_pct,
        "ev_pct": ev_pct,
        "rolling_accuracy_recent_pct": recent_acc * 100 if recent_acc is not None else None,
        "drift_window": drift_window,
        "drift_alert": drift_alert,
        "feature_importance": feature_importance,
    }


# ----------------------------
# 今日收盤方向機率預估（開盤後估計，樣本外回測）
# ----------------------------
#
# 跟「明日」模型的關鍵差異：這裡預測的是「今天」自己的 Open→Close 方向，
# 所以絕對不能用今天的收盤價、今天算出來的 RSI/MACD 等任何用到「今天收盤」
# 的資訊當特徵——那等於用答案預測答案，準確率會被灌水到失真。
# 只能用「今天開盤當下就已經確定」的資訊：今天的隔夜跳空（開盤 vs 昨收），
# 以及昨天（含）之前就已經算好、不會再變動的技術/量價特徵。

INTRADAY_FEATURE_COLS = [
    "overnight_gap_today", "intraday_return_prev", "volatility_5d_prev", "volatility_20d_prev",
    "rsi_diff_prev", "macd_hist_slope_prev", "volume_ratio_prev", "mom5_prev",
    "regime_trending_prev", "regime_ranging_prev",
]


def _build_intraday_features(df, intermarket_feat=None, intermarket_cols=None):
    """
    建立「今天開盤後」可用的特徵：今天的隔夜跳空(今天才知道) + 昨天(以前)
    已經確定、不會再變動的其他特徵(用 shift(1) 取得，避免任何當日收盤資訊)。

    intermarket_feat/intermarket_cols：跟 _build_direction_features 一樣選用，
    但這裡一律用 shift(1)（前一天已確定的值），因為今天大盤指標的當日收盤
    在開盤當下同樣還不知道，不能直接使用今天的值。
    """
    base = _build_direction_features(df, intermarket_feat=intermarket_feat)  # 每列都是「用到當天收盤」算出來的
    prev = base.shift(1)
    feat = pd.DataFrame(index=df.index)
    feat["overnight_gap_today"] = base["overnight_gap"]  # 今天開盤 vs 昨收，開盤當下就知道
    feat["intraday_return_prev"] = prev["intraday_return"]
    feat["volatility_5d_prev"] = prev["volatility_5d"]
    feat["volatility_20d_prev"] = prev["volatility_20d"]
    feat["rsi_diff_prev"] = prev["rsi_diff"]
    feat["macd_hist_slope_prev"] = prev["macd_hist_slope"]
    feat["volume_ratio_prev"] = prev["volume_ratio"]
    feat["mom5_prev"] = prev["mom5"]
    feat["regime_trending_prev"] = prev["regime_trending"]
    feat["regime_ranging_prev"] = prev["regime_ranging"]
    for c in (intermarket_cols or []):
        feat[f"{c}_prev"] = prev[c]
    return feat


def intraday_close_probability_walkforward(df, train_window=250, retrain_every=10, model="logistic",
                                            intermarket_feat=None, intermarket_cols=None,
                                            last_bar_complete=True):
    """
    估計「今天收盤是否高於今天開盤」的機率，只用開盤當下已知的資訊，
    用 walk-forward 回測誠實回報樣本外準確率(不是訓練集準確率)。
    回傳 None 代表資料量不足，或指定的模型無法使用。

    last_bar_complete (v3.7 新增)：傳入的 df 最後一根K棒是否已收盤。
    盤中執行時傳 False，此時最後一列的 target=(Close>Open) 是用未完成的
    Close 算出來的，不能拿來當答案，因此會從 walk-forward 的準確率統計中
    排除（訓練集本來就已排除，見下方標籤洩漏修正）。

    註：本模型的「特徵」本身不受未完成K棒影響——它只用今天的開盤價
    (overnight_gap_today) 與昨日(含)以前 shift(1) 的值，從不碰今天的 Close。
    所以盤中執行時應該傳入「含當日K棒」的完整序列，才拿得到今天的開盤價；
    這一點與隔日模型相反（隔日模型必須剔除未完成K棒）。
    """
    if model == "random_forest" and not _SKLEARN_AVAILABLE:
        return None

    df = calc_rsi(df.copy())
    df = calc_macd(df)
    feat = _build_intraday_features(df, intermarket_feat=intermarket_feat, intermarket_cols=intermarket_cols)
    feature_cols = INTRADAY_FEATURE_COLS + [f"{c}_prev" for c in (intermarket_cols or [])]
    feat_valid = feat.dropna(subset=feature_cols)

    target = (df["Close"] > df["Open"]).astype(int)  # 當天自己的 Open→Close 方向
    bt_data = feat_valid.join(target.rename("target")).dropna()

    if len(bt_data) < train_window + 30:
        return None

    X_all = bt_data[feature_cols].values
    y_all = bt_data["target"].values
    n = len(bt_data)

    # v3.7 修正（標籤洩漏）：本函式要預測的是 feat_valid 最後一列（今天）的
    # Open→Close 方向，但 target=(Close>Open) 對今天這一列也算得出來，
    # 所以 bt_data 的最後一列就是「被預測的那一列」，而且帶著答案。
    # v3.6 的最終訓練用 X_all[-train_window:]，等於讓模型先看過今天的答案
    # 再預測今天。隔日模型因為 target=shift(-1) 會被 dropna() 砍掉末列，
    # 沒有這個問題——兩者原本不一致，這裡補齊。
    predict_index = feat_valid.index[-1]
    predict_pos = None
    if len(bt_data) and bt_data.index[-1] == predict_index:
        predict_pos = n - 1

    fit_predict = _fit_predict_random_forest if model == "random_forest" else _fit_predict_logistic

    # 盤中執行時，最後一列的答案是用未完成的收盤價算的，不列入準確率統計。
    backtest_end = n
    if not last_bar_complete and predict_pos is not None:
        backtest_end = predict_pos
    if backtest_end - train_window < 30:
        return None

    preds, actuals = [], []
    p_up_cache = None
    for t in range(train_window, backtest_end):
        if p_up_cache is None or (t - train_window) % retrain_every == 0:
            X_train = X_all[t - train_window:t]
            y_train = y_all[t - train_window:t]
            end = min(t + retrain_every, backtest_end)
            p_up_cache = fit_predict(X_train, y_train, X_all[t:end])
            cache_start = t
        p_up = p_up_cache[t - cache_start]
        preds.append(int(p_up >= 0.5))
        actuals.append(int(y_all[t]))

    preds = np.array(preds)
    actuals = np.array(actuals)
    accuracy = float((preds == actuals).mean())
    always_up_acc = float((actuals == 1).mean())
    naive_baseline = max(always_up_acc, 1 - always_up_acc)

    # 用「今天」自己這一列的特徵(今天開盤當下已知)重新訓練、預測今天。
    # 訓練集嚴格排除被預測的那一列本身（見上方 v3.7 修正說明）。
    train_end = predict_pos if predict_pos is not None else n
    if train_end < train_window + 1:
        return None
    X_train_final = X_all[train_end - train_window:train_end]
    y_train_final = y_all[train_end - train_window:train_end]
    latest_feat = feat_valid[feature_cols].iloc[-1].values.reshape(1, -1)
    p_up_today = float(fit_predict(X_train_final, y_train_final, latest_feat)[0])

    return {
        "model": model,
        "p_up_today": p_up_today,
        "walk_forward_accuracy": accuracy * 100,
        "naive_baseline_accuracy": naive_baseline * 100,
        "edge_over_baseline": (accuracy - naive_baseline) * 100,
        "sample_size": len(preds),
        # 被預測的那一天是否已經收盤（答案其實已知）。收盤後執行時，
        # 這個模型講的是「今天開盤到收盤」——那件事已經發生完了，
        # 機率值只有回顧意義，不是可交易的預測。
        "target_date": predict_index.date() if hasattr(predict_index, "date") else None,
        "result_already_known": bool(last_bar_complete),
    }


# ----------------------------
# 長期記錄：匯出到固定 Excel 檔案
# ----------------------------

VERSION = "3.7"

EXCEL_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_analysis_log.xlsx")

# Excel 儲存格文字上限約 32,767 字元；保留緩衝，超過此長度就截斷。
FULL_OUTPUT_MAX_CHARS = 30000

EXCEL_LOG_COLUMNS = [
    "執行時間", "股票代碼", "公司名稱", "股價日期(資料基準日)", "預測目標日(隔日估計)", "目前股價", "幣別",
    # v3.7 新增：執行情境。沒有這幾欄就無法事後區分「盤中跑的」與「收盤後跑的」，
    # 而兩者的可信度差很多（見 _trim_incomplete_bar 的實測證據）。
    "是否盤中執行", "已剔除未完成K棒", "執行當下價格", "資料落後天數", "資料是否停滯",
    "較同批次落後天數", "是否較同批次落後", "跳過原因", "交易日曆提醒",
    "本益比", "預估本益比", "ROE(%)", "淨利率(%)", "營收成長率YoY(%)", "獲利成長率YoY(%)",
    "殖利率(%)", "殖利率是否異常", "負債權益比", "負債權益比是否異常",
    "產業(Sector)", "細分產業(Industry)", "SOX視為半導體產業週期代理",
    "基本面分數", "技術面分數", "籌碼面分數", "綜合分數", "綜合結論",
    "市場狀態", "ADX",
    "日波動度std(%)", "ATR", "統計期望值價格", "68%區間下緣", "68%區間上緣", "95%區間下緣", "95%區間上緣",
    "今日模型是否已知結果",
    "今日_邏輯迴歸_收高於開盤機率(%)", "今日_邏輯迴歸_樣本外準確率(%)", "今日_邏輯迴歸_優勢(pp)",
    "今日_邏輯迴歸_樣本數",
    "今日_RF_收高於開盤機率(%)", "今日_RF_樣本外準確率(%)", "今日_RF_優勢(pp)", "今日_RF_樣本數",
    "隔日_邏輯迴歸_上漲機率(%)", "隔日_邏輯迴歸_樣本外準確率(%)", "隔日_邏輯迴歸_優勢(pp)",
    "隔日_邏輯迴歸_勝率(%)", "隔日_邏輯迴歸_平均獲利(%)", "隔日_邏輯迴歸_平均虧損(%)", "隔日_邏輯迴歸_期望值EV(%)",
    "隔日_邏輯迴歸_近30日滾動準確率(%)", "隔日_邏輯迴歸_衰退警報", "隔日_邏輯迴歸_係數絕對值Top5",
    # v3.7 新增：樣本數與信賴下限。walk-forward 一直有算 sample_size，
    # 但 v3.6 沒有記錄，導致無法判斷任何一筆準確率是幾筆樣本算出來的。
    "隔日_邏輯迴歸_樣本數", "隔日_邏輯迴歸_準確率信賴下限(%)",
    "隔日_RF_上漲機率(%)", "隔日_RF_樣本外準確率(%)", "隔日_RF_優勢(pp)",
    "隔日_RF_勝率(%)", "隔日_RF_平均獲利(%)", "隔日_RF_平均虧損(%)", "隔日_RF_期望值EV(%)",
    "隔日_RF_近30日滾動準確率(%)", "隔日_RF_衰退警報", "隔日_RF_特徵重要性Top5",
    "隔日_RF_樣本數", "隔日_RF_準確率信賴下限(%)",
    "Strategy_Decision", "Strategy_Decision_信心(%)", "model_quality",
    # v3.7 新增：EV 影子決策，與現行規則並行記錄供日後比較
    "EV_Decision", "EV_採用模型", "EV_淨期望值(%)", "EV_來回成本(%)", "EV_原因",
    "Risk_Reward_Ratio", "建議停損價", "建議目標價",
    "籌碼資料日期", "外資最新日買賣超(股)", "外資5日斜率",
    "三大法人連續同方向天數", "三大法人連續方向", "外資買賣超佔當日成交量比重(%)", "外資佔量比重5日斜率",
    "程式版本",
    # v3.7 新增：供 tools/log_review.py 事後回填實際結果。
    # v3.6 的 75 個欄位裡沒有任何一個記錄「後來到底漲還是跌」，
    # 所以跑再多筆也無法回答「這套系統準不準」——這是最關鍵的缺口。
    "實際目標日收盤", "實際報酬(%)", "實際方向", "是否命中_邏輯迴歸", "是否命中_RF",
    "是否命中_綜合分數", "回填時間",
    "完整終端輸出",
]


class _TeeOutput:
    """
    把 sys.stdout 包一層：畫面上照原樣顯示(不影響使用者體驗)，
    同時把同一份文字複製一份到記憶體緩衝區，供分析結束後整段寫入 Excel。
    """

    def __init__(self, real_stdout):
        self.real_stdout = real_stdout
        self.buffer = io.StringIO()

    def write(self, s):
        self.real_stdout.write(s)
        self.buffer.write(s)
        return len(s)

    def flush(self):
        self.real_stdout.flush()

    def getvalue(self):
        return self.buffer.getvalue()

    def __getattr__(self, name):
        # 委派給真正的 stdout：像 yfinance/requests/urllib3 這類套件，
        # 有時會檢查 .isatty()、.encoding 等屬性；_TeeOutput 本身沒有
        # 實作這些，如果不轉發，遇到這種檢查會直接丟 AttributeError，
        # 導致分析中途崩潰、來不及寫入 Excel。
        return getattr(self.real_stdout, name)


def _truncate_for_excel(text, max_chars=FULL_OUTPUT_MAX_CHARS):
    if text is None:
        return ""
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...(內容過長，已截斷；完整輸出請見終端機畫面)"


def _append_to_excel_log(record, path=None):
    """
    把這次分析結果新增一列到固定的 Excel 檔案，供長期記錄比對。

    v3.7：path 預設改為 None、在函式內才解析成 EXCEL_LOG_PATH。原本寫成
    `path=EXCEL_LOG_PATH` 的預設值在函式「定義當下」就綁死了，之後改
    模組層級的 EXCEL_LOG_PATH 不會生效（測試要指定暫存路徑時會踩到）。

    設計原則：
    - 每次執行都是「新增一列」，不覆蓋、不刪除舊紀錄，也不自作主張幫你去重——
      同一天跑很多次，Excel 就會有很多列，濾重與取捨交給你自己在 Excel 裡處理，
      不由程式片面決定「哪一筆才算數」。
    - 如果偵測到現有檔案的欄位跟目前版本的 schema 不一致（例如用舊版本程式
      建立的紀錄檔），不會硬寫進去把資料弄亂，而是另外開一個新檔名，並在畫面上
      明確告訴你，讓你自己決定要不要手動合併。
    """
    if path is None:
        path = EXCEL_LOG_PATH
    if not _OPENPYXL_AVAILABLE:
        print("  ⚠ 未安裝 openpyxl，無法寫入 Excel 紀錄檔 (pip install openpyxl)；本次分析結果不會被記錄。")
        return False

    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            existing_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))] if ws.max_row >= 1 else []
            if existing_header != EXCEL_LOG_COLUMNS:
                fallback_path = path.replace(".xlsx", f"_v{VERSION}.xlsx")
                print(f"  ⚠ 既有 Excel 紀錄檔（{path}）欄位與目前版本(v{VERSION})不一致，"
                      f"可能是舊版本程式建立的檔案。為避免破壞既有資料，"
                      f"本次改寫入新檔案: {fallback_path}（請自行決定是否手動合併兩份檔案）")
                path = fallback_path
                if os.path.exists(path):
                    wb = load_workbook(path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "分析紀錄"
                    ws.append(EXCEL_LOG_COLUMNS)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "分析紀錄"
            ws.append(EXCEL_LOG_COLUMNS)

        row = [record.get(col, "") for col in EXCEL_LOG_COLUMNS]
        ws.append(row)
        wb.save(path)
        print(f"  ✓ 已寫入長期記錄: {path}")
        return True
    except PermissionError:
        print(f"  ⚠ 無法寫入 {path}（檔案可能正在 Excel 中開啟，請先關閉檔案再執行）")
        return _append_csv_fallback(record, path, "Excel 檔案被鎖定")
    except Exception as e:
        print(f"  ⚠ 寫入 Excel 紀錄檔失敗: {type(e).__name__}: {e}")
        return _append_csv_fallback(record, path, f"{type(e).__name__}: {e}")


def _append_csv_fallback(record, excel_path, reason):
    """
    v3.7 新增：Excel 寫入失敗時改存 CSV，不讓這筆分析結果直接消失。
    最常見的觸發情境就是你正好把 stock_analysis_log.xlsx 開在 Excel 裡。
    CSV 之後可以自行合併回 Excel。
    """
    import csv
    csv_path = excel_path.replace(".xlsx", "_fallback.csv")
    try:
        exists = os.path.exists(csv_path)
        with open(csv_path, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=EXCEL_LOG_COLUMNS, extrasaction="ignore")
            if not exists:
                writer.writeheader()
            writer.writerow({c: record.get(c, "") for c in EXCEL_LOG_COLUMNS})
        print(f"  ✓ 已改存 CSV 備援: {csv_path}（原因: {reason}）")
        return True
    except Exception as e2:
        print(f"  ⚠ CSV 備援也失敗，本次結果未被記錄: {type(e2).__name__}: {e2}")
        return False


def _append_skip_record(ticker_symbol, decision, reason, full_output=""):
    """
    v3.7 新增：資料不足或資料停滯而沒有實際分析時，仍留下一筆明確標記的
    紀錄，但 Strategy_Decision 記為 Skipped_*，不會被誤讀成「分析後認為
    中性」。這是為了修掉 020020.TW 那種情況——市場狀態 unknown、所有
    機率欄位都是 NaN，卻給了綜合分數 0 與「中性（訊號混合）」的結論。
    """
    record = {
        "執行時間": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "股票代碼": ticker_symbol,
        "Strategy_Decision": decision,
        "跳過原因": reason,
        "model_quality": "unknown",
        "程式版本": VERSION,
        "完整終端輸出": _truncate_for_excel(full_output),
    }
    _append_to_excel_log(record)


# ----------------------------
# 策略決策層：雙模型過濾器 + 風報比
# ----------------------------

def dual_model_decision(p_logistic, p_rf, threshold=0.58):
    """
    雙模型一致性過濾器：只有邏輯迴歸與 Random Forest 都存在、預測方向一致、
    且雙方信心（機率偏離 50% 的程度）都達到 threshold 時，才輸出明確的
    Buy/Sell；其餘情況一律 Wait。

    這是刻意保守的設計：「不交易」本身也是一種合理的策略選擇，比起在
    模型意見分歧或信心不足時仍勉強輸出方向，Wait 更誠實地反映目前的
    不確定性。threshold=0.58 是使用者指定的預設值，不是從資料 fit 出來的
    最適參數，可自行調整。
    """
    if p_logistic is None or p_rf is None:
        return "Wait", None, "雙模型未同時可用（資料量不足或未安裝 scikit-learn）"

    dir_logistic = p_logistic >= 0.5
    dir_rf = p_rf >= 0.5
    if dir_logistic != dir_rf:
        return "Wait", None, "雙模型方向不一致"

    conf_logistic = max(p_logistic, 1 - p_logistic)
    conf_rf = max(p_rf, 1 - p_rf)
    min_conf = min(conf_logistic, conf_rf)
    if min_conf < threshold:
        return "Wait", min_conf, f"信心未達門檻（{min_conf*100:.1f}% < {threshold*100:.0f}%）"

    decision = "Buy" if dir_logistic else "Sell"
    return decision, min_conf, f"雙模型同向且信心皆 ≥ {threshold*100:.0f}%"


def compute_risk_reward(current_price, atr, decision, atr_multiplier=1.5, reward_multiplier=2.0):
    """
    僅在有明確 Buy/Sell 訊號時才有意義。停損 = 進場價 ∓ atr_multiplier×ATR，
    目標價 = 進場價 ± reward_multiplier×ATR（粗略估計，非精確獲利目標），
    風報比 = 目標與進場價差 / 停損與進場價差。
    """
    if decision not in ("Buy", "Sell") or atr is None or pd.isna(atr) or current_price is None:
        return None
    if decision == "Buy":
        stop = current_price - atr_multiplier * atr
        target = current_price + reward_multiplier * atr
    else:
        stop = current_price + atr_multiplier * atr
        target = current_price - reward_multiplier * atr
    risk = abs(current_price - stop)
    reward = abs(target - current_price)
    rr_ratio = reward / risk if risk > 0 else None
    return {"stop_loss": stop, "target": target, "risk_reward_ratio": rr_ratio}


# ----------------------------
# 主流程
# ----------------------------

def analyze(ticker_symbol):
    _tee = _TeeOutput(sys.stdout)
    _old_stdout = sys.stdout
    sys.stdout = _tee
    try:
        print(f"\n{'='*50}")
        print(f"  分析標的: {ticker_symbol}")
        print(f"{'='*50}\n")

        ticker_obj = yf.Ticker(ticker_symbol)

        hist_raw = ticker_obj.history(period="3y", auto_adjust=True)
        if hist_raw.empty:
            print("查無此股票代碼的資料,請確認代碼是否正確"
                  "(台股請加 .TW 或 .TWO,如 2330.TW)")
            _append_skip_record(ticker_symbol, "Skipped_資料不足",
                                "查無此股票代碼的價格資料", _tee.getvalue())
            return

        # ---- v3.7 資料品質把關（三道關卡，任何一道不過就不輸出交易決策）----

        # 關卡一：盤中執行時剔除尚未收盤的K棒。詳見 _trim_incomplete_bar()。
        # 注意 hist_raw 要保留：今日(Open→Close)模型需要今天的開盤價，
        # 而它的特徵從不碰今天的收盤價，所以不受未完成K棒影響。
        hist, bar_trimmed, live_price = _trim_incomplete_bar(hist_raw, ticker_symbol)
        if bar_trimmed:
            print(f"⚠ 盤中執行偵測：已剔除今日尚未收盤的K棒（當下價格 {live_price:.2f}），")
            print("   改以前一個完整交易日為資料基準日。盤中的收盤價尚未定案，")
            print("   直接拿來算特徵會讓模型在沒見過的輸入上外插（v3.7 修正）。\n")
        if hist.empty:
            print("剔除未完成K棒後已無可用資料。")
            _append_skip_record(ticker_symbol, "Skipped_資料不足",
                                "剔除未完成K棒後無可用資料", _tee.getvalue())
            return

        # 關卡二：資料新鮮度。抓不到新資料時要說「抓不到」，不能拿舊資料當今天用。
        freshness = _check_price_freshness(hist)
        if freshness["stale"]:
            print(f"⚠ {freshness['message']}")
            print("   常見原因：該標的下市/暫停交易、代碼變更，或資料源當日未更新。")
            print("   本次不輸出交易決策，僅記錄狀態。\n")
            _append_skip_record(ticker_symbol, "Skipped_資料停滯",
                                freshness["message"], _tee.getvalue())
            return

        # 關卡三：資料量是否足以支撐分析。「資料不足」與「分析後認為中性」
        # 是兩件不同的事，不該長得一樣。
        gate = _should_analyze(hist)
        if not gate["ok"]:
            print(f"⚠ 資料不足以進行有意義的分析：{gate['reason']}")
            print("   本次不輸出綜合分數與交易決策，避免與「分析後認為中性」混淆。\n")
            _append_skip_record(ticker_symbol, "Skipped_資料不足",
                                gate["reason"], _tee.getvalue())
            return

        hist = calc_ma(hist)
        hist = calc_rsi(hist)
        hist = calc_macd(hist)
        hist = calc_kd(hist)
        hist = calc_bollinger(hist)
        hist = calc_adx(hist)

        current_price = float(hist["Close"].iloc[-1])
        regime, adx_value = classify_market_regime(hist)
        tech_notes, tech_score = technical_summary(hist, regime=regime, adx_value=adx_value)
        fund_notes, fund_score, info, fund_metrics = fundamental_summary(ticker_obj, current_price=current_price)
        chip_notes, chip_score, chip_date_raw, chip_metrics = chip_summary(ticker_symbol, ticker_obj, info=info)

        company_name = info.get("longName", ticker_symbol)
        currency = info.get("currency", "")
        price_date = hist.index[-1].date()

        print(f"公司名稱: {company_name}")
        if bar_trimmed:
            print(f"基準日收盤價: {current_price:.2f} {currency}"
                  f"（資料基準日 {price_date}；執行當下盤中價格 {live_price:.2f}）")
            print("  ※ 以下所有分析都以基準日收盤價為準，不使用盤中未定案的價格。\n")
        else:
            print(f"目前股價: {current_price:.2f} {currency}（資料基準日 {price_date}）\n")

        # 資料新鮮度檢查：籌碼面資料日期若與股價最新交易日相差過多（>3 個曆日，
        # 涵蓋週末），提示可能是抓取失敗或延遲，避免誤把舊籌碼資料當成最新狀況解讀。
        chip_date = _parse_chip_date(chip_date_raw)
        if chip_date is not None:
            gap_days = (price_date - chip_date).days
            if gap_days > 3:
                print(f"⚠ 資料新鮮度提醒: 籌碼面資料日期({chip_date})與最新股價日期({price_date})"
                      f"相差 {gap_days} 天，籌碼資料可能非最新，解讀時請留意\n")

        # 與同批次其他標的比較是否相對落後（絕對天數門檻抓不到的情況）
        batch_lagged, batch_lag_days, batch_lag_msg = _check_batch_lag(price_date)
        if batch_lagged:
            print(f"⚠ 批次資料落後提醒：{batch_lag_msg}\n")

        print("--- 基本面分析 ---")
        for n in fund_notes:
            print(f"  • {n}")
        print(f"  基本面分數: {fund_score}\n")

        print("--- 技術面分析 ---")
        for n in tech_notes:
            print(f"  • {n}")
        print(f"  技術面分數: {tech_score}\n")

        print("--- 籌碼面分析 ---")
        for n in chip_notes:
            print(f"  • {n}")
        # 外資買賣超佔當日成交量比重：用籌碼資料日期去 hist 找當日成交量比對，
        # 找不到對應交易日(例如抓取失敗或日期對不齊)就略過，不硬湊數字。
        foreign_ratio_pct = None
        concentration_slope = None
        if chip_metrics.get("foreign_net_latest") is not None and chip_date is not None:
            matched_vol = None
            for idx in hist.index:
                if idx.date() == chip_date:
                    matched_vol = float(hist.loc[idx, "Volume"])
                    break
            if matched_vol and matched_vol > 0:
                foreign_ratio_pct = chip_metrics["foreign_net_latest"] / matched_vol * 100
                print(f"  • 外資買賣超佔當日成交量比重: {foreign_ratio_pct:+.2f}%")
                # 報告層級指標(比照外資5日斜率的定位，不是模型訓練特徵——
                # 原因見 _append_concentration_cache 的說明)：本地累積後才能看趨勢。
                code_for_cache = ticker_symbol.upper().replace(".TW", "").replace(".TWO", "")
                _append_concentration_cache(code_for_cache, chip_date.strftime("%Y%m%d"), foreign_ratio_pct)
                concentration_slope, conc_n = _concentration_slope_from_cache(code_for_cache, lookback=5)
                if concentration_slope is not None:
                    direction = "轉強" if concentration_slope > 0 else "轉弱" if concentration_slope < 0 else "持平"
                    print(f"  • 外資佔量比重 5 日斜率(本地快取): {concentration_slope:+.2f} 個百分點/交易日（{direction}）")
                else:
                    print(f"  • 外資佔量比重斜率: 本地快取僅累積 {conc_n} 個交易日，需累積滿5個交易日"
                          "（即連續執行本程式5個不同交易日）才會開始計算，這是預期行為")
        print(f"  籌碼面分數: {chip_score}\n")

        # 大盤關聯指標(費城半導體/美元兌台幣/美國10年期公債殖利率)特徵：
        # 同一次執行只抓一次，批次查詢多檔股票時共用同一份快取。
        intermarket_feat, intermarket_cols = get_intermarket_features(hist.index)
        if not intermarket_cols:
            print("⚠ 大盤關聯指標(SOX/美元兌台幣/美債殖利率)本次皆抓取失敗，方向機率模型將只使用個股自身特徵。\n")

        # 產業週期：目前沒有可靠、免費的「產業循環階段」資料源可以量化，
        # 因此不硬造一個週期因子分數。能做到的是：若判斷為半導體相關產業，
        # 明確標註 SOX(費城半導體指數) 特徵在此扮演「半導體產業週期代理指標」
        # 的角色(SOX 走勢本身反映全球半導體業景氣)，讓你知道這個關聯性存在，
        # 而不是假裝有一個更精確的權重公式。其他產業目前沒有對應的代理指標。
        sector_str = (fund_metrics.get("sector") or "")
        industry_str = (fund_metrics.get("industry") or "")
        is_semiconductor = "semiconductor" in industry_str.lower() or "semiconductor" in sector_str.lower()
        if is_semiconductor and "SOX_dev20" in intermarket_cols:
            print(f"  • 產業別偵測為半導體相關({industry_str or sector_str})："
                  "SOX(費城半導體指數)特徵在此視為半導體產業週期代理指標，"
                  "並非精確的週期階段判斷，僅供參考關聯性。\n")

        total = fund_score + tech_score + chip_score
        print("--- 綜合結論 ---")
        # 分數區間對照表：讓「3 分代表什麼」有明確依據，而不是只有正負判斷。
        # 依實際加減分規則逐項核算（非拍腦袋估計）：
        #   基本面：ROE+淨利率(+1/-1) + 營收獲利同向(+1/-1) + 負債比過高(-1，只扣不加) → 區間 [-3, +2]
        #   技術面：均線排列(±2) + MACD(±1) + KD交叉(±1) + 量價(±1) → 區間 [-5, +5]
        #   籌碼面：台股多日法人(±1)；美股僅內部人買超弱訊號(0/+1，不會扣分) → 區間 [-1, +1]（美股上緣為 [0,+1]）
        # 故台股總分理論區間約為 -9 ~ +8（非對稱），美股略窄。下方 verdict 門檻是
        # 相對這個區間的粗略切分，score_to_probability() 那類「機率」映射不應直接套用此分數。
        if total >= 4:
            verdict = "偏多（訊號一致度較高，不等於買進建議）"
        elif total >= 2:
            verdict = "中性偏多"
        elif total <= -4:
            verdict = "偏空（訊號一致度較高，不等於賣出建議）"
        elif total <= -2:
            verdict = "中性偏空"
        else:
            verdict = "中性（訊號混合，各面向未形成一致方向）"
        print(f"  綜合分數: {total} → {verdict}")
        print(f"  （子分數: 基本面 {fund_score} / 技術面 {tech_score} / 籌碼面 {chip_score}；"
              f"參考區間 -6~+8，非對稱分布，分數愈極端代表三面向訊號愈一致，"
              f"分數本身不代表報酬率或勝率）")

        print("\n--- 隔日歷史經驗波動區間（非預測值）---")
        rng = next_day_range(hist)
        if rng is not None:
            print(f"  最近 {rng['window']} 日 Close-to-Close 報酬標準差: {rng['daily_std_pct']:.2f}%")
            print(f"  最近 {rng['window']} 日平均日報酬: {rng['mean_return_pct']:.2f}%")
            print(f"  ATR(14日 Wilder 平滑): {rng['atr']:.2f} {currency}")
            print(f"  統計期望值(歷史報酬中位數推算): {rng['median_price']:.2f} {currency}")
            print(f"  歷史經驗約 68% 區間(16%~84%分位): {rng['range_68'][0]:.2f} ~ {rng['range_68'][1]:.2f} {currency}")
            print(f"  歷史經驗約 95% 區間(2.5%~97.5%分位): {rng['range_95'][0]:.2f} ~ {rng['range_95'][1]:.2f} {currency}")
            print("  ※ 中位數推算值是「歷史報酬分布的中心點」，不是預測值；報酬分布本身高度離散，")
            print("     單一數字沒有實際下單意義，務必與上方區間一起看。")
            print("  ※ 改用歷史報酬分位數，不假設 Normal distribution；仍只描述 Close-to-Close，無法涵蓋盤中 High/Low 與所有跳空風險。")

        bt = backtest_confidence(hist)
        if bt is not None:
            print("\n  【歷史經驗區間覆蓋率回測】")
            print(f"  回測樣本數: {bt['sample_size']} 個交易日")
            print(f"  目標約 68% 區間，實際命中率: {bt['empirical_conf_68']:.1f}%（差異 {bt['target_68_gap']:+.1f} 個百分點）")
            print(f"  目標約 95% 區間，實際命中率: {bt['empirical_conf_95']:.1f}%（差異 {bt['target_95_gap']:+.1f} 個百分點）")
            print("  ※ 每一天只使用當天之前的資料建立區間，避免 look-ahead bias。")
            print("  ※ 覆蓋率接近目標不代表股價可預測；只代表區間校準程度較好。")

        target_date, calendar_note = next_trading_day_estimate(price_date)
        if calendar_note:
            print(f"\n⚠ 交易日曆提醒：{calendar_note}")

        intraday_label = ("今日" if bar_trimmed else f"{price_date} 當日")
        print(f"\n--- {intraday_label}收盤方向機率預估（開盤後估計，樣本外回測）---")
        if not bar_trimmed:
            print(f"  ※ 本次於盤前或收盤後執行，{price_date} 這一天的開盤到收盤已經結束，")
            print("     下方機率只有回顧意義，不是可交易的預測。這個模型要拿來當")
            print("     即時判斷用，必須在該日盤中執行。")

        def _print_intraday_result(dp, model_label):
            if dp is None:
                return
            direction_label = "收高於開盤" if dp["p_up_today"] >= 0.5 else "收低於開盤"
            print(f"  [{model_label}] {price_date} 「{direction_label}」機率: "
                  f"{max(dp['p_up_today'], 1 - dp['p_up_today']) * 100:.1f}%"
                  f"（收高於開盤機率原始值: {dp['p_up_today']*100:.1f}%）")
            print(f"  [{model_label}] 樣本外準確率: {dp['walk_forward_accuracy']:.1f}%"
                  f" / baseline: {dp['naive_baseline_accuracy']:.1f}%"
                  f" / 優勢: {dp['edge_over_baseline']:+.1f} 個百分點"
                  f"（樣本數 {dp['sample_size']}）")

        # v3.7：今日模型改用「含當日K棒」的 hist_raw——它需要今天的開盤價，
        # 而它的特徵從不碰今天的收盤價（見 _build_intraday_features）。
        # last_bar_complete=not bar_trimmed 告訴它最後一列的答案能不能採信。
        intraday_source = hist_raw if bar_trimmed else hist
        intraday_feat, intraday_cols = (
            get_intermarket_features(intraday_source.index)
            if bar_trimmed else (intermarket_feat, intermarket_cols)
        )
        ip_logistic = intraday_close_probability_walkforward(
            intraday_source, model="logistic", intermarket_feat=intraday_feat,
            intermarket_cols=intraday_cols, last_bar_complete=not bar_trimmed)
        ip_rf = intraday_close_probability_walkforward(
            intraday_source, model="random_forest", intermarket_feat=intraday_feat,
            intermarket_cols=intraday_cols, last_bar_complete=not bar_trimmed)

        if ip_logistic is None and ip_rf is None:
            print("  資料量不足以支撐 walk-forward 回測（需要至少約 280 個交易日），略過此區塊。")
        else:
            _print_intraday_result(ip_logistic, "邏輯迴歸")
            if ip_rf is not None:
                _print_intraday_result(ip_rf, "Random Forest")
            elif not _SKLEARN_AVAILABLE:
                print("  [Random Forest] 未安裝 scikit-learn，略過此模型")
            print("  ※ 這裡預測的是「今天開盤後」到「今天收盤」這段區間，只用開盤當下已知的")
            print("     資訊(隔夜跳空+前一交易日已確定的技術/量價特徵)，不使用今天盤中任何")
            print("     時間點的價格——本工具沒有歷史盤中時間戳資料，無法誠實回測「現在幾點、")
            print("     股價在哪」這種條件下的機率，所以不做這種宣稱。")

        print(f"\n--- 隔日方向機率預估（資料基準日: {price_date} → 預測目標日: {target_date} 估計，"
              f"已跳過週末與已知國定假日，樣本外回測，非保證預測）---")

        def _print_direction_result(dp, model_label):
            if dp is None:
                return
            direction_label = "上漲" if dp["p_up_tomorrow"] >= 0.5 else "下跌"
            print(f"  [{model_label}] {target_date} 「{direction_label}」機率: "
                  f"{max(dp['p_up_tomorrow'], 1 - dp['p_up_tomorrow']) * 100:.1f}%"
                  f"（上漲機率原始值: {dp['p_up_tomorrow']*100:.1f}%）")
            print(f"  [{model_label}] 樣本外準確率: {dp['walk_forward_accuracy']:.1f}%"
                  f" / baseline: {dp['naive_baseline_accuracy']:.1f}%"
                  f" / 優勢: {dp['edge_over_baseline']:+.1f} 個百分點"
                  f"（樣本數 {dp['sample_size']}）")
            if dp["avg_gain_pct"] is not None and dp["avg_loss_pct"] is not None:
                print(f"  [{model_label}] 期望值(EV，若照方向交易，不計手續費/稅/滑價): "
                      f"勝率 {dp['win_rate_pct']:.1f}% / 平均獲利 {dp['avg_gain_pct']:+.2f}% "
                      f"/ 平均虧損 -{dp['avg_loss_pct']:.2f}% / EV {dp['ev_pct']:+.3f}%")
            else:
                print(f"  [{model_label}] 期望值(EV): 回測樣本中缺少獲利或虧損其中一類交易，無法計算")
            if dp["rolling_accuracy_recent_pct"] is not None:
                print(f"  [{model_label}] 近{dp['drift_window']}個交易日滾動樣本外準確率: "
                      f"{dp['rolling_accuracy_recent_pct']:.1f}%")
                if dp["drift_alert"]:
                    print(f"  ⚠ [{model_label}] 模型衰退警報：近{dp['drift_window']}個交易日樣本外準確率低於50%，"
                          "建議考慮暫停使用此模型判斷、檢查特徵是否仍然有效，或重新訓練。")
            if dp.get("feature_importance"):
                top5 = sorted(dp["feature_importance"].items(), key=lambda kv: kv[1], reverse=True)[:5]
                top5_str = "、".join(f"{k}={v:.3f}" for k, v in top5)
                if dp["model"] == "random_forest":
                    print(f"  [{model_label}] 特徵重要性 Top5(Gini重要性): {top5_str}")
                else:
                    print(f"  [{model_label}] 標準化係數絕對值 Top5(非嚴謹重要性，僅供排序參考): {top5_str}")

        dp_logistic = direction_probability_walkforward(
            hist, model="logistic", intermarket_feat=intermarket_feat, intermarket_cols=intermarket_cols)
        dp_rf = direction_probability_walkforward(
            hist, model="random_forest", intermarket_feat=intermarket_feat, intermarket_cols=intermarket_cols)

        if dp_logistic is None and dp_rf is None:
            print("  資料量不足以支撐 walk-forward 回測（需要至少約 280 個交易日），略過此區塊。")
        else:
            _print_direction_result(dp_logistic, "邏輯迴歸")
            if dp_rf is not None:
                _print_direction_result(dp_rf, "Random Forest")
            elif not _SKLEARN_AVAILABLE:
                print("  [Random Forest] 未安裝 scikit-learn，略過此模型（pip install scikit-learn 後可啟用對照組）")

            edges = [dp["edge_over_baseline"] for dp in (dp_logistic, dp_rf) if dp is not None]
            if edges and max(edges) < 2.0:
                print("  ⚠ 所有模型的準確率都沒有明顯超過 naive baseline，代表目前這組特徵")
                print("     幾乎沒有提供超額判斷力，機率數字僅供參考，不建議作為進出場依據。")
            if dp_logistic is not None and dp_rf is not None:
                better = "Random Forest" if dp_rf["edge_over_baseline"] > dp_logistic["edge_over_baseline"] else "邏輯迴歸"
                print(f"  ※ 本次樣本外回測中，{better} 的優勢分數較高；但單一標的、單一時間窗格的比較")
                print("     樣本數有限，這個「誰比較準」的結論不能直接套用到其他股票或其他時間段。")
            print("  ※ 這是機率估計，不是保證預測。即使準確率有 55~60%，仍代表有相當高機率判斷錯誤方向；")
            print("     機率輸出本身也未經 reliability diagram / Brier score 等校準度檢驗，僅代表方向判斷的粗略傾向。")

        # 雙模型一致性過濾器 + Strategy_Decision + Risk_Reward_Ratio
        p_logistic_tomorrow = dp_logistic["p_up_tomorrow"] if dp_logistic is not None else None
        p_rf_tomorrow = dp_rf["p_up_tomorrow"] if dp_rf is not None else None

        # 模型品質關卡（v3.7 改版）：這是比「單次預測信心」更上位的檢查——
        # 不管這次雙模型信心多高、方向多一致，只要模型本身的優勢無法與隨機
        # 區分，再強的單次信心也不可信，一律強制 Wait。
        #
        # v3.6 用的是「樣本外準確率 < 55% 就封鎖」，問題是這個 55% 是個
        # 與樣本數無關的固定值，兩個方向都會判錯：
        #   - n=60 時 56% 可以通關，但那個區間寬到蓋住 50%，根本分不出優勢
        #   - n=2000 時 53.5% 被擋下，但那已經是統計上站得住腳的結果
        #
        # v3.7 改用 Wilson 信賴區間下限：要求「連下限都高於 50%」，
        # 門檻隨樣本數自動調整（n=60 需 ≥62.7%、n=250 需 ≥56.2%、
        # n=1000 需 ≥53.1%）。
        #
        # 附帶一提：實測中 2891C.TW 回報 74~77% 準確率（全場最高，也是
        # 唯一觸發訊號的一檔），而其餘標的中位數只有 51~52%。那 74% 究竟
        # 是真優勢還是資料問題，在 v3.6 下根本無從判斷——因為 sample_size
        # 算了卻沒被記錄。這道關卡本身不保證能擋掉它（若樣本數確實夠多，
        # 74% 是會通過的，而且應該通過）；真正該擋的是它的成因，也就是
        # 上面 _trim_incomplete_bar() 修掉的未完成K棒問題。這裡要修的是
        # 另一件事：讓「幾筆樣本算出來的」這個資訊不再遺失。
        rel_logistic = assess_accuracy_reliability(
            dp_logistic["walk_forward_accuracy"], dp_logistic["sample_size"]
        ) if dp_logistic is not None else None
        rel_rf = assess_accuracy_reliability(
            dp_rf["walk_forward_accuracy"], dp_rf["sample_size"]
        ) if dp_rf is not None else None

        unreliable_models = []
        if rel_logistic is not None and not rel_logistic["reliable"]:
            unreliable_models.append(f"邏輯迴歸（{rel_logistic['message']}）")
        if rel_rf is not None and not rel_rf["reliable"]:
            unreliable_models.append(f"Random Forest（{rel_rf['message']}）")

        if unreliable_models:
            strategy_decision, decision_conf = "Wait", None
            decision_reason = ("模型品質關卡：" + "；".join(unreliable_models)
                               + "。不論雙模型信心是否達標，一律封鎖買賣建議")
        else:
            strategy_decision, decision_conf, decision_reason = dual_model_decision(p_logistic_tomorrow, p_rf_tomorrow)
        rr = compute_risk_reward(current_price, rng["atr"] if rng is not None else None, strategy_decision)

        # model_quality：給下游(Excel/dashboard)用的機器可讀標籤，判斷邏輯
        # 跟上面的模型品質關卡共用同一套 Wilson 判準，維持一致。
        def _classify_model_quality(rel_a, rel_b, dp_a, dp_b):
            rels = [r for r in (rel_a, rel_b) if r is not None]
            edges = [dp["edge_over_baseline"] for dp in (dp_a, dp_b) if dp is not None]
            if not rels:
                return "unknown"
            if not all(r["reliable"] for r in rels):
                # 兩者取較差的標籤：weak 代表點估計好看但信賴區間仍蓋到 50%
                return "weak" if any(r["label"] == "weak" for r in rels) else "not_reliable"
            if edges and max(edges) < 2.0:
                return "weak"
            return "usable_with_caution"

        model_quality = _classify_model_quality(rel_logistic, rel_rf, dp_logistic, dp_rf)

        # ---- EV 影子決策（v3.7 新增，不取代上面的 Strategy_Decision）----
        # 問的不是「兩個模型是不是都超過 58%」，而是「扣掉台股來回成本之後
        # 期望值還是正的嗎」。刻意採影子模式：在累積足夠樣本證明哪一種
        # 規則較好之前，不片面改掉你原本的交易邏輯，只把兩者都記進 Excel。
        ev_source = None
        for _dp, _rel in ((dp_logistic, rel_logistic), (dp_rf, rel_rf)):
            if _dp is None or _rel is None:
                continue
            if ev_source is None or _dp["edge_over_baseline"] > ev_source[0]["edge_over_baseline"]:
                ev_source = (_dp, _rel)
        if ev_source is not None:
            _dp, _rel = ev_source
            ev_decision, ev_detail = expectancy_decision(
                _dp["p_up_tomorrow"], _dp["avg_gain_pct"], _dp["avg_loss_pct"], _rel)
            ev_model_label = "邏輯迴歸" if _dp["model"] == "logistic" else "Random Forest"
        else:
            ev_decision, ev_detail, ev_model_label = "Wait", {"cost_pct": round_trip_cost_pct(),
                                                              "ev_pct": None,
                                                              "reason": "沒有可用的方向模型"}, None

        print(f"\n--- Strategy_Decision（雙模型過濾器：信心門檻 58% + Wilson 信賴下限 > 50%）---")
        conf_text = f"{decision_conf*100:.1f}%" if decision_conf is not None else "N/A"
        print(f"  Strategy_Decision: {strategy_decision}（信心: {conf_text}；原因: {decision_reason}）")
        print(f"  model_quality: {model_quality}")
        for _label, _rel in (("邏輯迴歸", rel_logistic), ("Random Forest", rel_rf)):
            if _rel is not None:
                print(f"    - [{_label}] {_rel['message']}")
        print(f"\n  【EV 影子決策（v3.7 新增，僅記錄比較用，不是交易指示）】")
        print(f"  EV_Decision: {ev_decision}"
              + (f"（採用 {ev_model_label}）" if ev_model_label else ""))
        print(f"  原因: {ev_detail['reason']}")
        print(f"  ※ 這是與現行規則並行記錄的第二種判斷方式。累積足夠樣本後，")
        print(f"     用 tools/log_review.py 比較兩者哪一個實際表現較好，再決定要不要換。")
        if rr is not None:
            print(f"  建議停損價: {rr['stop_loss']:.2f} {currency} / 建議目標價: {rr['target']:.2f} {currency}"
                  f"（皆為進場價 ± ATR 倍數的粗略估計，非精確停利停損點）")
            if rr["risk_reward_ratio"] is not None:
                print(f"  Risk_Reward_Ratio: {rr['risk_reward_ratio']:.2f}")
        print("  ※ Wait 不代表「看空」或「沒有機會」，只代表這次雙模型判斷不夠一致、信心不足，")
        print("     或模型整體準確率未達品質門檻，刻意不勉強輸出方向；")
        print("     「不交易」本身也是一種合理的策略選擇。")

        print("\n--- 方法論限制 ---")
        print("  • 基本面比率主要來自 yfinance 資料快照；嚴格估值仍應使用同業/自身歷史分位數。")
        print("  • 技術分數是規則式描述，不代表上漲機率；尚未用 OOS 資料校準成勝率。")
        print("  • 籌碼面對台股使用多日法人流量，避免單日買賣超直接等同多空。")
        print("  • 波動區間為 Close-to-Close 歷史經驗分布；若要模擬實際下單，下一版應再拆解 Overnight Gap、Open-to-Close、High/Low。")
        print("  • 本程式仍不是完整交易策略回測器；尚未計入手續費、交易稅、滑價、成交限制與流動性。")
        print("\n※ 以上分析僅為歷史數據統計整理，不代表未來走勢預測，亦不構成投資建議。")

        # ------------------------------------------------------------
        # 組裝長期記錄用的 record，寫入固定 Excel 檔案。
        # 全部取自上面已經算好的結構化數值（fund_metrics/chip_metrics/rng/bt/dp_*），
        # 不重新解析畫面文字，確保 Excel 裡的數字跟你剛才看到的輸出是同一份資料。
        # ------------------------------------------------------------
        # 取得到目前為止(record 組裝前)完整的終端輸出文字，供寫入 Excel。
        # 刻意在這裡取值，不含稍後 _append_to_excel_log 內部印出的
        # 「已寫入長期記錄」確認訊息（避免自我指涉）。
        # ------------------------------------------------------------
        _full_output_text = _truncate_for_excel(_tee.getvalue())

        record = {
            "執行時間": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "股票代碼": ticker_symbol,
            "公司名稱": company_name,
            "股價日期(資料基準日)": str(price_date),
            "預測目標日(隔日估計)": str(target_date),
            "目前股價": current_price,
            "幣別": currency,
            # v3.7 新增：執行情境。事後統計時務必只採用「是否盤中執行=False」
            # 的紀錄——盤中跑的那些不可信（見 _trim_incomplete_bar 的實測證據）。
            "是否盤中執行": bool(bar_trimmed),
            "已剔除未完成K棒": bool(bar_trimmed),
            "執行當下價格": live_price,
            "資料落後天數": freshness.get("lag_days"),
            "資料是否停滯": bool(freshness.get("stale")),
            "較同批次落後天數": batch_lag_days,
            "是否較同批次落後": bool(batch_lagged),
            "交易日曆提醒": calendar_note or None,
            "本益比": fund_metrics.get("pe"),
            "預估本益比": fund_metrics.get("forward_pe"),
            "ROE(%)": fund_metrics.get("roe_pct"),
            "淨利率(%)": fund_metrics.get("profit_margin_pct"),
            "營收成長率YoY(%)": fund_metrics.get("rev_growth_pct"),
            "獲利成長率YoY(%)": fund_metrics.get("earnings_growth_pct"),
            "殖利率(%)": fund_metrics.get("dividend_yield_pct"),
            "殖利率是否異常": fund_metrics.get("dividend_yield_abnormal"),
            "負債權益比": fund_metrics.get("debt_to_equity"),
            "負債權益比是否異常": fund_metrics.get("debt_to_equity_abnormal"),
            "產業(Sector)": fund_metrics.get("sector"),
            "細分產業(Industry)": fund_metrics.get("industry"),
            "SOX視為半導體產業週期代理": is_semiconductor and "SOX_dev20" in intermarket_cols,
            "基本面分數": fund_score,
            "技術面分數": tech_score,
            "籌碼面分數": chip_score,
            "綜合分數": total,
            "綜合結論": verdict,
            "市場狀態": regime,
            "ADX": adx_value,
            "程式版本": VERSION,
            "完整終端輸出": _full_output_text,
        }
        if rng is not None:
            record.update({
                "日波動度std(%)": rng["daily_std_pct"],
                "ATR": rng["atr"],
                "統計期望值價格": rng["median_price"],
                "68%區間下緣": rng["range_68"][0],
                "68%區間上緣": rng["range_68"][1],
                "95%區間下緣": rng["range_95"][0],
                "95%區間上緣": rng["range_95"][1],
            })
        # 今日模型：收盤後執行時，被預測的那一天其實已經結束，機率只有回顧
        # 意義。這個旗標讓事後統計能把「回顧」與「即時預測」分開看。
        record["今日模型是否已知結果"] = not bool(bar_trimmed)
        if ip_logistic is not None:
            record.update({
                "今日_邏輯迴歸_收高於開盤機率(%)": ip_logistic["p_up_today"] * 100,
                "今日_邏輯迴歸_樣本外準確率(%)": ip_logistic["walk_forward_accuracy"],
                "今日_邏輯迴歸_優勢(pp)": ip_logistic["edge_over_baseline"],
                "今日_邏輯迴歸_樣本數": ip_logistic["sample_size"],
            })
        if ip_rf is not None:
            record.update({
                "今日_RF_收高於開盤機率(%)": ip_rf["p_up_today"] * 100,
                "今日_RF_樣本外準確率(%)": ip_rf["walk_forward_accuracy"],
                "今日_RF_優勢(pp)": ip_rf["edge_over_baseline"],
                "今日_RF_樣本數": ip_rf["sample_size"],
            })
        if dp_logistic is not None:
            record.update({
                "隔日_邏輯迴歸_上漲機率(%)": dp_logistic["p_up_tomorrow"] * 100,
                "隔日_邏輯迴歸_樣本外準確率(%)": dp_logistic["walk_forward_accuracy"],
                "隔日_邏輯迴歸_優勢(pp)": dp_logistic["edge_over_baseline"],
                "隔日_邏輯迴歸_勝率(%)": dp_logistic["win_rate_pct"],
                "隔日_邏輯迴歸_平均獲利(%)": dp_logistic["avg_gain_pct"],
                "隔日_邏輯迴歸_平均虧損(%)": dp_logistic["avg_loss_pct"],
                "隔日_邏輯迴歸_期望值EV(%)": dp_logistic["ev_pct"],
                "隔日_邏輯迴歸_近30日滾動準確率(%)": dp_logistic["rolling_accuracy_recent_pct"],
                "隔日_邏輯迴歸_衰退警報": dp_logistic["drift_alert"],
                "隔日_邏輯迴歸_樣本數": dp_logistic["sample_size"],
                "隔日_邏輯迴歸_準確率信賴下限(%)": (
                    rel_logistic["lower_bound"] if rel_logistic else None),
            })
            if dp_logistic.get("feature_importance"):
                top5 = sorted(dp_logistic["feature_importance"].items(), key=lambda kv: kv[1], reverse=True)[:5]
                record["隔日_邏輯迴歸_係數絕對值Top5"] = "、".join(f"{k}={v:.3f}" for k, v in top5)
        if dp_rf is not None:
            record.update({
                "隔日_RF_上漲機率(%)": dp_rf["p_up_tomorrow"] * 100,
                "隔日_RF_樣本外準確率(%)": dp_rf["walk_forward_accuracy"],
                "隔日_RF_優勢(pp)": dp_rf["edge_over_baseline"],
                "隔日_RF_勝率(%)": dp_rf["win_rate_pct"],
                "隔日_RF_平均獲利(%)": dp_rf["avg_gain_pct"],
                "隔日_RF_平均虧損(%)": dp_rf["avg_loss_pct"],
                "隔日_RF_期望值EV(%)": dp_rf["ev_pct"],
                "隔日_RF_近30日滾動準確率(%)": dp_rf["rolling_accuracy_recent_pct"],
                "隔日_RF_衰退警報": dp_rf["drift_alert"],
                "隔日_RF_樣本數": dp_rf["sample_size"],
                "隔日_RF_準確率信賴下限(%)": rel_rf["lower_bound"] if rel_rf else None,
            })
            if dp_rf.get("feature_importance"):
                top5 = sorted(dp_rf["feature_importance"].items(), key=lambda kv: kv[1], reverse=True)[:5]
                record["隔日_RF_特徵重要性Top5"] = "、".join(f"{k}={v:.3f}" for k, v in top5)
        record["Strategy_Decision"] = strategy_decision
        record["Strategy_Decision_信心(%)"] = decision_conf * 100 if decision_conf is not None else None
        record["model_quality"] = model_quality
        record["EV_Decision"] = ev_decision
        record["EV_採用模型"] = ev_model_label
        record["EV_淨期望值(%)"] = ev_detail.get("ev_pct")
        record["EV_來回成本(%)"] = ev_detail.get("cost_pct")
        record["EV_原因"] = ev_detail.get("reason")
        if rr is not None:
            record["Risk_Reward_Ratio"] = rr["risk_reward_ratio"]
            record["建議停損價"] = rr["stop_loss"]
            record["建議目標價"] = rr["target"]
        if chip_metrics.get("net_streak_days") is not None:
            record["三大法人連續同方向天數"] = chip_metrics["net_streak_days"]
            record["三大法人連續方向"] = chip_metrics.get("net_streak_direction")
        if foreign_ratio_pct is not None:
            record["外資買賣超佔當日成交量比重(%)"] = foreign_ratio_pct
        if concentration_slope is not None:
            record["外資佔量比重5日斜率"] = concentration_slope
        if chip_date_raw:
            record["籌碼資料日期"] = str(chip_date_raw)
        if chip_metrics.get("foreign_net_latest") is not None:
            record["外資最新日買賣超(股)"] = chip_metrics["foreign_net_latest"]
        if chip_metrics.get("foreign_5d_slope") is not None:
            record["外資5日斜率"] = chip_metrics["foreign_5d_slope"]

        _append_to_excel_log(record)
    finally:
        sys.stdout = _old_stdout


def score_to_probability(raw_score, beta=0.5):
    """
    將綜合分數透過 Sigmoid 映射到 (0,1)。

    ⚠ 統計警告（刻意保留、不要刪除這段提醒）：
    這不是一個經過驗證的機率估計。beta=0.5 是任意設定的常數，
    不是從歷史資料 fit 出來的參數；raw_score 本身也只是規則式加總分數，
    沒有證據顯示它與「未來上漲」存在線性且穩定的關係。
    在沒有做以下事情之前，這個函式的輸出不能當成真實勝率使用：
      1. 用歷史資料建立 (raw_score -> 隔日/隔週報酬方向) 的樣本
      2. 用 Logistic Regression 或其他方式 fit 出 beta（而非手動指定）
      3. 用 Out-of-Sample 資料驗證校準度（例如 reliability diagram /
         Brier score），確認「模型說 70% 的時候，實際真的接近 70%」
    目前這個函式僅供展示 Sigmoid 映射的形狀，不建議在報告中直接
    當作「上漲機率」呈現給使用者，以免誤導。
    """
    probability = 1 / (1 + np.exp(-beta * raw_score))
    return probability


def _run_batch(symbols, pause_sec=1.0):
    """
    逐檔執行 analyze()，並在每檔之間加入短暫延遲——
    多檔查詢會讓台股籌碼面(TWSE T86)的請求量倍增，這道延遲是為了
    降低短時間內被 TWSE 判定為異常流量、暫時封鎖的風險。
    單一股票查詢失敗（例如代碼打錯）不應該讓後面排隊的股票也查不到，
    因此個別 analyze() 出錯會被攔截、記錄，並繼續跑下一檔。

    v3.6 補強：過濾掉空白/空字串代碼(避免指令列多打一個空格就整個中斷)，
    結尾加上成功/失敗筆數統計。
    """
    global _BATCH_LATEST_DATE
    _BATCH_LATEST_DATE = None   # 每次批次重新開始比較基準

    clean_symbols = [s.strip() for s in symbols if s and s.strip()]
    if not clean_symbols:
        print("⚠ 沒有有效的股票代碼可供分析。")
        return

    results_summary = []
    total = len(clean_symbols)
    success_count = 0
    fail_count = 0
    for idx, symbol in enumerate(clean_symbols, 1):
        print(f"\n[{idx}/{total}] 開始分析 {symbol} ...")
        try:
            analyze(symbol)
            results_summary.append((symbol, "完成"))
            success_count += 1
        except Exception as e:
            print(f"⚠ 分析 {symbol} 時發生錯誤，略過此檔繼續下一檔: {type(e).__name__}: {e}")
            results_summary.append((symbol, f"失敗（{type(e).__name__}）"))
            fail_count += 1
        if idx < total and pause_sec > 0:
            time.sleep(pause_sec)  # 多檔之間的緩衝，降低對 TWSE 的短時間請求壓力

    if total > 1:
        print(f"\n{'='*50}")
        print("  批次查詢結果總覽")
        print(f"{'='*50}")
        for symbol, status in results_summary:
            print(f"  {symbol}: {status}")
        print(f"\n  成功: {success_count} / 失敗: {fail_count} / 總數: {total}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        _prog = os.path.basename(sys.argv[0])  # 動態取得實際檔名,升版改檔名不會再漏改這裡
        print(f"用法: python {_prog} <股票代碼> [股票代碼2] [股票代碼3] ...")
        print(f"範例(單檔): python {_prog} 2330.TW")
        print(f"範例(多檔): python {_prog} 2330.TW 2317.TW AAPL")
        sys.exit(1)

    _run_batch(sys.argv[1:])